"""Unit tests for the Sharesight sync skill."""

from __future__ import annotations

from datetime import date, datetime, time
import os
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from openclaw.skills.sharesight_sync.skill import (
    PayoutRecord,
    PortfolioRecord,
    SharesightApiClient,
    SharesightSyncSkill,
)


def create_workbook_fixture(workbook_path: Path, *, duplicate_pay_date: bool = False) -> None:
    """Create a workbook fixture with the DC Pavula worksheet layout."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DC Pavula FY2526"

    headers = [
        "Pay Date",
        "Dist / Share",
        "ROC%",
        "Inc%",
        "ROC $",
        "Inc $",
        "Gross Amt",
        "Tax",
        "GA-Tax",
        "Excess Tax",
        "Act Tax",
        "FXRates",
    ]
    worksheet.cell(row=6, column=8).value = "MSTY Distributions USD 2025/2026"
    for offset, header in enumerate(headers, start=8):
        worksheet.cell(row=7, column=offset).value = header

    rows = [
        (8, datetime(2026, 1, 5), 75.5, 123.45, 130.23, 6.78, 2.34, 1.2345),
        (9, datetime(2026, 1, 12), 35.21, 56.78, 60.0, 1.11, 0.22, 1.1111),
    ]
    if duplicate_pay_date:
        rows.append((10, datetime(2026, 1, 5), 88.88, 99.99, 111.11, 2.22, 0.33, 1.2222))

    for row_index, pay_date, inc_percent, inc_amount, gross_amount, tax, excess_tax, fx_rate in rows:
        worksheet.cell(row=row_index, column=8).value = pay_date
        worksheet.cell(row=row_index, column=11).value = inc_percent
        worksheet.cell(row=row_index, column=13).value = inc_amount
        worksheet.cell(row=row_index, column=14).value = gross_amount
        worksheet.cell(row=row_index, column=15).value = tax
        worksheet.cell(row=row_index, column=17).value = excess_tax
        worksheet.cell(row=row_index, column=20).value = fx_rate

    workbook.save(workbook_path)
    workbook.close()


class FakeApiClient:
    """A test double for the Sharesight API client."""

    def __init__(
        self,
        payouts: list[PayoutRecord],
        *,
        portfolio: PortfolioRecord | None = None,
        payout_fallbacks: dict[int, PayoutRecord] | None = None,
        confirmed_payouts: dict[str, PayoutRecord] | None = None,
    ) -> None:
        self.portfolio = portfolio or PortfolioRecord(id=1201759, name="DC Pavula")
        self.payouts = payouts
        self.payout_fallbacks = payout_fallbacks or {}
        self.confirmed_payouts = confirmed_payouts or {}
        self.closed = False
        self.resolve_portfolio_calls: list[str] = []
        self.list_payouts_calls: list[tuple[int, date | None, date | None]] = []
        self.get_payout_calls: list[int] = []
        self.confirm_payout_calls: list[dict[str, object]] = []
        self.updated_payloads: list[tuple[int, dict[str, object]]] = []

    def resolve_portfolio(self, portfolio_name: str) -> PortfolioRecord:
        """Return the configured fake portfolio."""
        self.resolve_portfolio_calls.append(portfolio_name)
        return self.portfolio

    def list_payouts(
        self,
        portfolio_id: int,
        *,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[PayoutRecord]:
        """Return the configured payout list."""
        self.list_payouts_calls.append((portfolio_id, start_date, end_date))
        return list(self.payouts)

    def get_payout(self, payout_id: int) -> PayoutRecord:
        """Return fallback payout detail when list data is incomplete."""
        self.get_payout_calls.append(payout_id)
        return self.payout_fallbacks[payout_id]

    def confirm_payout(self, payload: dict[str, object]) -> PayoutRecord:
        """Record payout confirmation and return the created payout."""
        self.confirm_payout_calls.append(payload)
        paid_on = payload["payout"]["paid_on"]
        return self.confirmed_payouts[str(paid_on)]

    def update_payout(self, payout_id: int, payload: dict[str, object]) -> dict[str, object]:
        """Record payout updates."""
        self.updated_payloads.append((payout_id, payload))
        return {"payout": {"id": payout_id}}

    def close(self) -> None:
        """Mark the fake client as closed."""
        self.closed = True


class TestSharesightSyncSkill(unittest.TestCase):
    """Unit coverage for the Sharesight sync skill."""

    def test_closest_portfolio_name_prefers_nearest_visible_match(self) -> None:
        """The API helper picks the closest visible candidate."""
        client = SharesightApiClient(
            api_base_url="https://api.sharesight.com/api/v2",
            token_url="https://api.sharesight.com/oauth/token",
            client_id="client-id",
            client_secret="client-secret",
        )
        self.assertEqual(
            client._closest_portfolio_name(
                "DC Pavula",
                ["DC Pavlov", "Growth Portfolio", "Pavula Family"],
            ),
            "DC Pavlov",
        )

    def test_constructor_stores_api_configuration(self) -> None:
        """The skill keeps configured API settings."""
        skill = SharesightSyncSkill(
            excel_path=Path("/tmp/Personal CashFlow.xlsx"),
            client_id="client-id",
            client_secret="client-secret",
            api_base_url="https://api.sharesight.com/api/v2",
            token_url="https://api.sharesight.com/oauth/token",
            tax_field_name="resident_withholding_tax",
            confirmed_state="confirmed",
            api_factory=lambda: FakeApiClient([]),
        )

        self.assertEqual(skill.api_base_url, "https://api.sharesight.com/api/v2")
        self.assertEqual(skill.token_url, "https://api.sharesight.com/oauth/token")
        self.assertEqual(skill.tax_field_name, "resident_withholding_tax")
        self.assertEqual(skill.confirmed_state, "confirmed")

    def test_constructor_uses_excel_path_from_environment_variable(self) -> None:
        """The skill can override the workbook path via environment variable."""
        original = os.environ.get("OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH")
        try:
            os.environ["OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH"] = "/tmp/live-cashflow.xlsx"
            skill = SharesightSyncSkill(
                client_id="client-id",
                client_secret="client-secret",
                excel_path_env_var="OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH",
                api_factory=lambda: FakeApiClient([]),
            )
        finally:
            if original is None:
                os.environ.pop("OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH", None)
            else:
                os.environ["OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH"] = original

        self.assertEqual(skill.excel_path, Path("/tmp/live-cashflow.xlsx"))

    def test_run_skips_worksheet_rows_with_excel_error_values(self) -> None:
        """Rows containing Excel error values are reported and ignored."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)

            workbook = load_workbook(workbook_path)
            worksheet = workbook["DC Pavula FY2526"]
            worksheet.cell(row=10, column=8).value = datetime(2026, 1, 19)
            worksheet.cell(row=10, column=11).value = 50
            worksheet.cell(row=10, column=13).value = 42
            worksheet.cell(row=10, column=14).value = 49
            worksheet.cell(row=10, column=15).value = 7
            worksheet.cell(row=10, column=17).value = 1
            worksheet.cell(row=10, column=20).value = "#N/A"
            workbook.save(workbook_path)
            workbook.close()

            api = FakeApiClient(
                [
                    PayoutRecord(id=1, paid_on=date(2026, 1, 5), state="unconfirmed", raw={}),
                    PayoutRecord(id=2, paid_on=date(2026, 1, 19), state="unconfirmed", raw={}),
                ],
            )
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["matched_pay_dates"], ["05/01/2026"])
        self.assertEqual(result["unmatched_pay_dates"], ["19/01/2026"])
        self.assertEqual(
            result["skipped_worksheet_rows"],
            [
                {
                    "row": "10",
                    "pay_date": "19/01/2026",
                    "reason": "Worksheet contains Excel error value: #N/A",
                },
            ],
        )

    def test_run_skips_rows_with_time_only_pay_date_values(self) -> None:
        """Rows with a time-only pay date are reported and ignored."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)

            workbook = load_workbook(workbook_path)
            worksheet = workbook["DC Pavula FY2526"]
            worksheet.cell(row=10, column=8).value = time(0, 0, 0)
            worksheet.cell(row=10, column=11).value = 50
            worksheet.cell(row=10, column=13).value = 42
            worksheet.cell(row=10, column=14).value = 49
            worksheet.cell(row=10, column=15).value = 7
            worksheet.cell(row=10, column=17).value = 1
            worksheet.cell(row=10, column=20).value = 1.25
            workbook.save(workbook_path)
            workbook.close()

            api = FakeApiClient(
                [PayoutRecord(id=1, paid_on=date(2026, 1, 5), state="unconfirmed", raw={})],
            )
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["matched_pay_dates"], ["05/01/2026"])
        self.assertEqual(result["unmatched_pay_dates"], [])
        self.assertEqual(
            result["skipped_worksheet_rows"],
            [
                {
                    "row": "10",
                    "pay_date": "00:00:00",
                    "reason": "Worksheet contains time-only value instead of date: 00:00:00",
                },
            ],
        )

    def test_run_dry_run_lists_matches_without_updating(self) -> None:
        """Dry-run mode reports matched pay dates and payloads without PUT requests."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient(
                [
                    PayoutRecord(id=10, paid_on=date(2026, 1, 5), state="unconfirmed", raw={}),
                    PayoutRecord(id=11, paid_on=date(2026, 1, 9), state="unconfirmed", raw={}),
                    PayoutRecord(id=12, paid_on=date(2026, 1, 12), state="unconfirmed", raw={}),
                ],
            )
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertTrue(result["dry_run"])
        self.assertEqual(result["unconfirmed_payouts_found"], 3)
        self.assertEqual(result["matched_and_updated"], 0)
        self.assertEqual(result["matched_pay_dates"], ["05/01/2026", "12/01/2026"])
        self.assertEqual(result["unmatched_pay_dates"], ["09/01/2026"])
        self.assertEqual(result["skipped_worksheet_rows"], [])
        self.assertEqual(api.updated_payloads, [])
        self.assertEqual(len(result["dry_run_payloads"]), 2)
        self.assertEqual(result["dry_run_payloads"][0]["payout_id"], 10)
        self.assertIsNone(result["dry_run_payloads"][0]["confirm_payload"])
        self.assertEqual(result["dry_run_payloads"][0]["update_payload"]["payout"]["amount"], 123.45)
        self.assertEqual(
            result["dry_run_payloads"][1]["update_payload"]["payout"]["exchange_rate"],
            1.1111,
        )

    def test_run_updates_only_exact_pay_date_matches(self) -> None:
        """Only exact worksheet pay dates are updated."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient(
                [
                    PayoutRecord(id=10, paid_on=date(2026, 1, 5), state="unconfirmed", raw={}),
                    PayoutRecord(id=11, paid_on=date(2026, 1, 9), state="unconfirmed", raw={}),
                    PayoutRecord(id=12, paid_on=date(2026, 1, 12), state="unconfirmed", raw={}),
                ],
            )
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["status"], "success")
        self.assertFalse(result["dry_run"])
        self.assertEqual(result["unconfirmed_payouts_found"], 3)
        self.assertEqual(result["matched_and_updated"], 2)
        self.assertEqual(result["matched_pay_dates"], ["05/01/2026", "12/01/2026"])
        self.assertEqual(result["unmatched_pay_dates"], ["09/01/2026"])
        self.assertEqual(result["skipped_worksheet_rows"], [])
        self.assertEqual(api.resolve_portfolio_calls, ["DC Pavula"])
        self.assertEqual(api.list_payouts_calls, [(1201759, date(2026, 1, 5), date(2026, 1, 12))])
        self.assertTrue(api.closed)

        first_payload = api.updated_payloads[0][1]["payout"]
        self.assertEqual(api.updated_payloads[0][0], 10)
        self.assertEqual(first_payload["paid_on"], "2026-01-05")
        self.assertEqual(first_payload["goes_ex_on"], "2026-01-04")
        self.assertEqual(first_payload["amount"], 123.45)
        self.assertEqual(first_payload["resident_withholding_tax"], 6.78)
        self.assertEqual(first_payload["exchange_rate"], 1.2345)
        self.assertEqual(
            first_payload["comments"],
            "MSTY Income: 75.5 of 130.23. 15% withholding includes 2.34 excess on ROC.",
        )
        self.assertEqual(api.updated_payloads[1][0], 12)
        self.assertEqual(api.updated_payloads[1][1]["payout"]["amount"], 56.78)

    def test_run_confirms_unconfirmed_payout_before_updating_when_id_missing(self) -> None:
        """Unconfirmed payouts without IDs are POSTed before the final PUT."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient(
                [
                    PayoutRecord(
                        id=None,
                        paid_on=date(2026, 1, 5),
                        state="unconfirmed",
                        raw={"holding_id": 25152995, "company_event_id": 13386665, "paid_on": "2026-01-05"},
                    ),
                ],
                confirmed_payouts={
                    "2026-01-05": PayoutRecord(
                        id=77,
                        paid_on=date(2026, 1, 5),
                        state="confirmed",
                        raw={"id": 77, "paid_on": "2026-01-05"},
                    ),
                },
            )
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["matched_and_updated"], 1)
        self.assertEqual(len(api.confirm_payout_calls), 1)
        self.assertEqual(
            api.confirm_payout_calls[0],
            {
                "payout": {
                    "holding_id": 25152995,
                    "company_event_id": 13386665,
                    "paid_on": "2026-01-05",
                    "state": "confirmed",
                },
            },
        )
        self.assertEqual(api.updated_payloads[0][0], 77)

    def test_run_uses_payout_show_when_list_data_lacks_paid_on(self) -> None:
        """Missing list dates fall back to the payout show call."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient(
                [PayoutRecord(id=20, paid_on=None, state="unconfirmed", raw={})],
                payout_fallbacks={
                    20: PayoutRecord(id=20, paid_on=date(2026, 1, 5), state="unconfirmed", raw={}),
                },
            )
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(api.get_payout_calls, [20])
        self.assertEqual(result["matched_pay_dates"], ["05/01/2026"])

    def test_parse_payout_record_allows_rows_without_ids(self) -> None:
        """Unconfirmed payout rows may omit IDs and must still be parsed."""
        client = SharesightApiClient(
            api_base_url="https://api.sharesight.com/api/v2",
            token_url="https://api.sharesight.com/oauth2/token",
            client_id="client-id",
            client_secret="client-secret",
        )

        rows = [
            {"id": None, "paid_on": "2026-01-05", "state": "unconfirmed"},
            {"id": 21, "paid_on": "2026-01-12", "state": "unconfirmed"},
        ]
        parsed = [client._parse_payout_record(row) for row in rows]

        self.assertIsNone(parsed[0].id)
        self.assertIsNotNone(parsed[1])
        self.assertEqual(parsed[1].id, 21)

    def test_run_raises_for_duplicate_pay_dates_in_worksheet(self) -> None:
        """Duplicate pay dates are rejected because matching must be exact."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path, duplicate_pay_date=True)
            skill = SharesightSyncSkill(
                excel_path=workbook_path,
                client_id="client-id",
                client_secret="client-secret",
                api_factory=lambda: FakeApiClient([]),
            )

            with self.assertRaisesRegex(ValueError, "duplicate Pay Date 05/01/2026"):
                skill.run()

    def test_constructor_requires_api_credentials_when_not_explicitly_supplied(self) -> None:
        """The skill raises a clear error when Sharesight API credentials are missing."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)

            with self.assertRaisesRegex(ValueError, "Sharesight client ID was not provided"):
                SharesightSyncSkill(
                    excel_path=workbook_path,
                    client_id_env="MISSING_SHARESIGHT_CLIENT_ID",
                    client_secret_env="MISSING_SHARESIGHT_CLIENT_SECRET",
                    api_factory=lambda: FakeApiClient([]),
                )


if __name__ == "__main__":
    unittest.main()
