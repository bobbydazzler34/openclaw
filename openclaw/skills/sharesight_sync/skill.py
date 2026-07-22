"""Sharesight sync skill implementation."""

from __future__ import annotations

from base64 import b64encode
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
import json
import os
from pathlib import Path
from typing import Any, Callable, Protocol
from urllib import error, parse, request

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from openclaw.skills._base.skill_base import SkillBase
from openclaw.skills.sharesight_sync.log_writer import ObsidianRunLogWriter

DEFAULT_API_BASE_URL = "https://api.sharesight.com/api/v2"
DEFAULT_TOKEN_URL = "https://api.sharesight.com/oauth2/token"
DEFAULT_EXCEL_PATH = "~/OneDrive/Documents/Finance/Personal CashFlow.xlsx"
DEFAULT_EXCEL_PATH_ENV_VAR = "OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH"
DEFAULT_WORKSHEET_NAME = "DC Pavula FY2627"
DEFAULT_PORTFOLIO_NAME = "DC Pavula"
DEFAULT_CLIENT_ID_ENV = "SHARESIGHT_CLIENT_ID"
DEFAULT_CLIENT_SECRET_ENV = "SHARESIGHT_CLIENT_SECRET"
DEFAULT_TAX_FIELD_NAME = "resident_withholding_tax"
DEFAULT_CONFIRMED_STATE = "confirmed"
DEFAULT_UNCONFIRMED_STATE = "unconfirmed"
DEFAULT_UPDATE_EXISTING_PAYOUTS_BY_ID = False
DEFAULT_LOG_ENV = "local"
DEFAULT_LOG_OPERATOR = "unknown"
DEFAULT_OBSIDIAN_USER = "bobbyd"
# Sharesight list_payouts end_date appears exclusive; buffer includes the worksheet max pay day.
PAYOUT_LIST_END_DATE_BUFFER_DAYS = 1
DATE_OUTPUT_FORMAT = "%d/%m/%Y"
HEADER_START_COLUMN = 8
REQUIRED_HEADERS = {
    "Pay Date": 8,
    "Inc%": 11,
    "Inc $": 13,
    "Gross Amt": 14,
    "Tax": 15,
    "Excess Tax": 17,
    "FXRates": 20,
}


@dataclass(slots=True)
class WorksheetEntry:
    """Normalized worksheet row keyed by pay date."""

    pay_date: date
    income_percent: str
    income_amount: str
    gross_amount: str
    foreign_tax_withheld: str
    excess_tax: str
    exchange_rate: str

    def to_payout_update(self, *, tax_field_name: str, confirmed_state: str) -> "PayoutUpdate":
        """Return the API payload for this worksheet row."""
        return PayoutUpdate(
            paid_on=self.pay_date,
            goes_ex_on=self.pay_date - timedelta(days=1),
            amount=self.income_amount,
            tax_field_name=tax_field_name,
            tax_amount=self.foreign_tax_withheld,
            exchange_rate=self.exchange_rate,
            comments=(
                f"MSTY Income: {self.income_percent} of {self.gross_amount}. "
                f"15% withholding includes {self.excess_tax} excess on ROC."
            ),
            state=confirmed_state,
        )


@dataclass(slots=True)
class PayoutUpdate:
    """Editable Sharesight payout fields."""

    paid_on: date
    goes_ex_on: date
    amount: str
    tax_field_name: str
    tax_amount: str
    exchange_rate: str
    comments: str
    state: str

    def to_api_payload(self, *, include_amount: bool = True) -> dict[str, Any]:
        """Return the wrapped Sharesight PUT payload."""
        payout_payload = {
            "paid_on": self.paid_on.isoformat(),
            "goes_ex_on": self.goes_ex_on.isoformat(),
            self.tax_field_name: float(self.tax_amount),
            "exchange_rate": float(self.exchange_rate),
            "comments": self.comments,
            "state": self.state,
        }
        if include_amount:
            payout_payload["amount"] = float(self.amount)
        return {"payout": payout_payload}


@dataclass(slots=True)
class PortfolioRecord:
    """A Sharesight portfolio summary."""

    id: int
    name: str


@dataclass(slots=True)
class PayoutRecord:
    """A Sharesight payout returned by the API."""

    id: int | None
    paid_on: date | None
    state: str
    raw: dict[str, Any]

    @property
    def holding_id(self) -> int | None:
        """Return the holding ID when present."""
        value = self.raw.get("holding_id")
        return None if value in (None, "") else int(value)

    @property
    def company_event_id(self) -> int | None:
        """Return the company event ID when present."""
        value = self.raw.get("company_event_id")
        return None if value in (None, "") else int(value)


class SharesightApi(Protocol):
    """API-facing contract for Sharesight interactions."""

    def resolve_portfolio(self, portfolio_name: str) -> PortfolioRecord:
        """Resolve the configured portfolio by exact or closest name."""

    def list_payouts(
        self,
        portfolio_id: int,
        *,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[PayoutRecord]:
        """Return payouts for the requested portfolio."""

    def get_payout(self, payout_id: int) -> PayoutRecord:
        """Return one payout, used as a fallback when list data is incomplete."""

    def confirm_payout(self, payload: dict[str, Any]) -> PayoutRecord:
        """Confirm or create an unconfirmed payout."""

    def update_payout(self, payout_id: int, payload: dict[str, Any]) -> dict[str, Any]:
        """Update one payout."""

    def close(self) -> None:
        """Close any API resources."""


@dataclass(slots=True)
class RunResult:
    """Structured outcome of a Sharesight sync run."""

    dry_run: bool
    portfolio_id: int
    portfolio_name: str
    unconfirmed_payouts_found: int
    matched_and_updated: int
    matched_pay_dates: list[str]
    unmatched_pay_dates: list[str]
    skipped_worksheet_rows: list[dict[str, str]]
    skipped_api_rows: list[dict[str, Any]]
    dry_run_payloads: list[dict[str, Any]]
    differing_income_tax_ids: list[int]


class SharesightSyncSkill(SkillBase):
    """Sync Sharesight unconfirmed payouts from the DC Pavula worksheet."""

    def __init__(
        self,
        config_path: str | Path | None = None,
        *,
        api_base_url: str | None = None,
        token_url: str | None = None,
        excel_path: str | Path | None = None,
        excel_path_env_var: str | None = None,
        worksheet_name: str | None = None,
        portfolio_name: str | None = None,
        client_id: str | None = None,
        client_secret: str | None = None,
        client_id_env: str | None = None,
        client_secret_env: str | None = None,
        dry_run: bool | None = None,
        tax_field_name: str | None = None,
        confirmed_state: str | None = None,
        unconfirmed_state: str | None = None,
        update_existing_payouts_by_id: bool | None = None,
        payouts_start_date: str | date | None = None,
        payouts_end_date: str | date | None = None,
        workbook_loader: Callable[..., Any] | None = None,
        api_factory: Callable[[], SharesightApi] | None = None,
    ) -> None:
        """Initialize the skill with optional dependency injection for testing."""
        super().__init__(config_path)
        self.api_base_url = str(self.config.get("api_base_url", api_base_url or DEFAULT_API_BASE_URL)).rstrip("/")
        self.token_url = str(self.config.get("token_url", token_url or DEFAULT_TOKEN_URL))
        self.excel_path_env_var = str(
            excel_path_env_var or self.config.get("excel_path_env_var", DEFAULT_EXCEL_PATH_ENV_VAR),
        )
        self.excel_path = self._resolve_excel_path(excel_path)
        self.worksheet_name = str(
            self.config.get("worksheet_name", worksheet_name or DEFAULT_WORKSHEET_NAME),
        )
        self.portfolio_name = str(
            self.config.get("portfolio_name", portfolio_name or DEFAULT_PORTFOLIO_NAME),
        )
        self.client_id = client_id or self._read_secret(
            env_name=str(self.config.get("client_id_env", client_id_env or DEFAULT_CLIENT_ID_ENV)),
            label="Sharesight client ID",
        )
        self.client_secret = client_secret or self._read_secret(
            env_name=str(self.config.get("client_secret_env", client_secret_env or DEFAULT_CLIENT_SECRET_ENV)),
            label="Sharesight client secret",
        )
        configured_dry_run = self._coerce_optional_bool(
            dry_run if dry_run is not None else self.config.get("dry_run"),
        )
        self.dry_run = False if configured_dry_run is None else configured_dry_run
        self.tax_field_name = str(
            self.config.get("tax_field_name", tax_field_name or DEFAULT_TAX_FIELD_NAME),
        )
        self.confirmed_state = str(
            self.config.get("confirmed_state", confirmed_state or DEFAULT_CONFIRMED_STATE),
        )
        self.unconfirmed_state = str(
            self.config.get("unconfirmed_state", unconfirmed_state or DEFAULT_UNCONFIRMED_STATE),
        )
        configured_update_existing = self._coerce_optional_bool(
            update_existing_payouts_by_id
            if update_existing_payouts_by_id is not None
            else self.config.get("update_existing_payouts_by_id"),
        )
        self.update_existing_payouts_by_id = (
            DEFAULT_UPDATE_EXISTING_PAYOUTS_BY_ID
            if configured_update_existing is None
            else configured_update_existing
        )
        self.payouts_start_date = self._coerce_optional_date(
            self.config.get("payouts_start_date", payouts_start_date),
        )
        self.payouts_end_date = self._coerce_optional_date(
            self.config.get("payouts_end_date", payouts_end_date),
        )
        self.workbook_loader = workbook_loader or load_workbook
        self._api_factory = api_factory or (
            lambda: SharesightApiClient(
                api_base_url=self.api_base_url,
                token_url=self.token_url,
                client_id=self.client_id,
                client_secret=self.client_secret,
            )
        )

    def _resolve_excel_path(self, excel_path: str | Path | None) -> Path:
        """Resolve the workbook path from explicit input, env var, or config."""
        if excel_path is not None:
            return Path(excel_path).expanduser()

        env_excel_path = os.getenv(self.excel_path_env_var)
        if env_excel_path:
            self.logger.info(
                "Using workbook path from environment variable %s.",
                self.excel_path_env_var,
            )
            return Path(env_excel_path).expanduser()

        configured_excel_path = self.config.get("excel_path", DEFAULT_EXCEL_PATH)
        return Path(str(configured_excel_path)).expanduser()

    def _write_obsidian_run_log(self, result_payload: dict[str, Any]) -> None:
        """Write an Obsidian run log when explicitly enabled in config."""
        if not bool(self.config.get("obsidian_log_enabled", False)):
            return

        configured_logs_dir = self.config.get("obsidian_log_dir")
        if not configured_logs_dir:
            self.logger.warning(
                "Obsidian logging enabled, but `obsidian_log_dir` is not configured.",
            )
            return

        log_writer = ObsidianRunLogWriter(
            logs_dir=Path(str(configured_logs_dir)).expanduser(),
            operator=str(self.config.get("obsidian_log_operator", DEFAULT_LOG_OPERATOR)),
            environment=str(self.config.get("obsidian_log_environment", DEFAULT_LOG_ENV)),
            obsidian_user=str(self.config.get("obsidian_log_user", DEFAULT_OBSIDIAN_USER)),
            config_path=str(
                self.config.get("obsidian_log_config_path", "skills/sharesight_sync/config.yaml"),
            ),
            env_var_name=str(self.config.get("excel_path_env_var", DEFAULT_EXCEL_PATH_ENV_VAR)),
        )

        output_path = log_writer.write_log(
            result_payload,
            excel_path_resolved=str(self.excel_path),
            env_override_used=bool(os.getenv(str(self.config.get("excel_path_env_var", DEFAULT_EXCEL_PATH_ENV_VAR)))),
            command_used=str(
                self.config.get(
                    "obsidian_log_command_used",
                    'python -c "from openclaw.skills.sharesight_sync.skill import SharesightSyncSkill; '
                    'print(SharesightSyncSkill(...).run())"',
                ),
            ),
            notes=["Generated automatically by SharesightSyncSkill."],
        )
        self.logger.info("Wrote Obsidian run log to %s.", output_path)

    def run(self) -> dict[str, Any]:
        """Update Sharesight unconfirmed payouts from the configured worksheet."""
        self.logger.info(
            "Starting Sharesight API sync for portfolio %s using %s [%s].",
            self.portfolio_name,
            self.excel_path,
            self.worksheet_name,
        )
        worksheet_entries, skipped_worksheet_rows = self._read_worksheet_entries()
        api = self._api_factory()
        matched_pay_dates: list[str] = []
        unmatched_pay_dates: list[str] = []
        skipped_api_rows: list[dict[str, Any]] = []
        dry_run_payloads: list[dict[str, Any]] = []
        differing_income_tax_ids: list[int] = []
        found_count = 0
        matched_count = 0

        try:
            portfolio = api.resolve_portfolio(self.portfolio_name)
            start_date, end_date, buffer_api_end = self._determine_payout_date_window(
                worksheet_entries,
            )
            api_end_date = self._sharesight_list_payouts_end_date(
                end_date,
                buffer_worksheet_end=buffer_api_end,
            )
            payouts = api.list_payouts(
                portfolio.id,
                start_date=start_date,
                end_date=api_end_date,
            )

            if self.update_existing_payouts_by_id:
                found_count, matched_count = self._run_update_existing_payouts_by_id(
                    payouts=payouts,
                    worksheet_entries=worksheet_entries,
                    api=api,
                    matched_pay_dates=matched_pay_dates,
                    unmatched_pay_dates=unmatched_pay_dates,
                    skipped_api_rows=skipped_api_rows,
                    dry_run_payloads=dry_run_payloads,
                    differing_income_tax_ids=differing_income_tax_ids,
                )
            else:
                found_count, matched_count = self._run_confirm_unconfirmed_flow(
                    payouts=payouts,
                    worksheet_entries=worksheet_entries,
                    api=api,
                    matched_pay_dates=matched_pay_dates,
                    unmatched_pay_dates=unmatched_pay_dates,
                    skipped_api_rows=skipped_api_rows,
                    dry_run_payloads=dry_run_payloads,
                )
        finally:
            api.close()

        result = RunResult(
            dry_run=self.dry_run,
            portfolio_id=portfolio.id,
            portfolio_name=portfolio.name,
            unconfirmed_payouts_found=found_count,
            matched_and_updated=matched_count,
            matched_pay_dates=matched_pay_dates,
            unmatched_pay_dates=unmatched_pay_dates,
            skipped_worksheet_rows=skipped_worksheet_rows,
            skipped_api_rows=skipped_api_rows,
            dry_run_payloads=dry_run_payloads,
            differing_income_tax_ids=differing_income_tax_ids,
        )
        self.logger.info(
            "Sharesight API sync complete with %s unconfirmed payouts and %s matches%s.",
            found_count,
            len(matched_pay_dates),
            " in dry-run mode" if self.dry_run else "",
        )
        result_payload = {
            "status": "success",
            "dry_run": result.dry_run,
            "portfolio_id": result.portfolio_id,
            "portfolio_name": result.portfolio_name,
            "api_base_url": self.api_base_url,
            "workbook_path": str(self.excel_path),
            "worksheet_name": self.worksheet_name,
            "tax_field_name": self.tax_field_name,
            "update_existing_payouts_by_id": self.update_existing_payouts_by_id,
            "confirmed_state": self.confirmed_state,
            "unconfirmed_state": self.unconfirmed_state,
            "payouts_start_date": None if start_date is None else start_date.isoformat(),
            "payouts_end_date": None if end_date is None else end_date.isoformat(),
            "payouts_api_end_date": None if api_end_date is None else api_end_date.isoformat(),
            "unconfirmed_payouts_found": result.unconfirmed_payouts_found,
            "matched_and_updated": result.matched_and_updated,
            "matched_pay_dates": result.matched_pay_dates,
            "unmatched_pay_dates": result.unmatched_pay_dates,
            "skipped_worksheet_rows": result.skipped_worksheet_rows,
            "skipped_api_rows": result.skipped_api_rows,
            "dry_run_payloads": result.dry_run_payloads,
            "differing_income_tax_ids": result.differing_income_tax_ids,
            "unconfirmed_transactions_found": result.unconfirmed_payouts_found,
        }
        self._write_obsidian_run_log(result_payload)
        return result_payload

    def _run_confirm_unconfirmed_flow(
        self,
        *,
        payouts: list[PayoutRecord],
        worksheet_entries: dict[date, WorksheetEntry],
        api: SharesightApi,
        matched_pay_dates: list[str],
        unmatched_pay_dates: list[str],
        skipped_api_rows: list[dict[str, Any]],
        dry_run_payloads: list[dict[str, Any]],
    ) -> tuple[int, int]:
        """Current flow: confirm unconfirmed payouts, then update them."""
        found_count = 0
        matched_count = 0
        for payout in payouts:
            if payout.state != self.unconfirmed_state:
                continue
            resolved_payout = payout
            if payout.id is not None and payout.paid_on is None:
                resolved_payout = api.get_payout(payout.id)
            if resolved_payout.paid_on is None:
                skipped_api_rows.append(
                    {
                        "reason": "Unconfirmed payout is missing paid_on.",
                        "payout": resolved_payout.raw,
                    },
                )
                continue
            found_count += 1
            entry = worksheet_entries.get(resolved_payout.paid_on)
            if entry is None:
                unmatched_pay_dates.append(self._format_date(resolved_payout.paid_on))
                continue
            matched_pay_dates.append(self._format_date(resolved_payout.paid_on))
            confirm_payload = None if resolved_payout.id is not None else self._build_confirm_payload(resolved_payout)
            if resolved_payout.id is None and confirm_payload is None:
                skipped_api_rows.append(
                    {
                        "reason": "Unconfirmed payout cannot be confirmed because holding_id or company_event_id is missing.",
                        "payout": resolved_payout.raw,
                    },
                )
                continue
            update = entry.to_payout_update(
                tax_field_name=self.tax_field_name,
                confirmed_state=self.confirmed_state,
            )
            include_amount = self._amount_should_be_included(update.amount)
            if not include_amount:
                self._log_omitted_amount(resolved_payout.id, resolved_payout.paid_on, update.amount)
            update_payload = update.to_api_payload(include_amount=include_amount)
            self._log_payout_update(
                resolved_payout.id,
                resolved_payout.paid_on,
                confirm_payload,
                update_payload,
            )
            if self.dry_run:
                dry_run_payloads.append(
                    {
                        "payout_id": resolved_payout.id,
                        "paid_on": resolved_payout.paid_on.isoformat(),
                        "confirm_payload": confirm_payload,
                        "update_payload": update_payload,
                    },
                )
                continue
            target_payout = resolved_payout
            if target_payout.id is None:
                target_payout = api.confirm_payout(confirm_payload)
            api.update_payout(target_payout.id, update_payload)
            matched_count += 1
        return found_count, matched_count

    def _run_update_existing_payouts_by_id(
        self,
        *,
        payouts: list[PayoutRecord],
        worksheet_entries: dict[date, WorksheetEntry],
        api: SharesightApi,
        matched_pay_dates: list[str],
        unmatched_pay_dates: list[str],
        skipped_api_rows: list[dict[str, Any]],
        dry_run_payloads: list[dict[str, Any]],
        differing_income_tax_ids: list[int],
    ) -> tuple[int, int]:
        """New flow: update existing payout IDs by paid_on match only."""
        found_count = 0
        matched_count = 0
        for payout in payouts:
            if payout.id is None:
                continue
            resolved_payout = payout if payout.paid_on is not None else api.get_payout(payout.id)
            if resolved_payout.paid_on is None:
                skipped_api_rows.append({"reason": "Payout is missing paid_on.", "payout": resolved_payout.raw})
                continue
            found_count += 1
            entry = worksheet_entries.get(resolved_payout.paid_on)
            if entry is None:
                unmatched_pay_dates.append(self._format_date(resolved_payout.paid_on))
                continue
            matched_pay_dates.append(self._format_date(resolved_payout.paid_on))
            update = entry.to_payout_update(
                tax_field_name=self.tax_field_name,
                confirmed_state=self.confirmed_state,
            )
            include_amount = self._amount_should_be_included(update.amount)
            if not include_amount:
                self._log_omitted_amount(resolved_payout.id, resolved_payout.paid_on, update.amount)
            update_payload = update.to_api_payload(include_amount=include_amount)
            current_amount = self._to_float_or_none(resolved_payout.raw.get("amount"))
            current_tax = self._to_float_or_none(resolved_payout.raw.get(self.tax_field_name))
            desired_amount = self._to_float_or_none(update_payload["payout"].get("amount"))
            desired_tax = update_payload["payout"][self.tax_field_name]
            if not self._floats_differ(current_amount, desired_amount) and not self._floats_differ(current_tax, desired_tax):
                continue
            differing_income_tax_ids.append(resolved_payout.id)
            if self.dry_run:
                dry_run_payloads.append(
                    {
                        "payout_id": resolved_payout.id,
                        "paid_on": resolved_payout.paid_on.isoformat(),
                        "current_amount": current_amount,
                        "desired_amount": desired_amount,
                        "current_tax": current_tax,
                        "desired_tax": desired_tax,
                        "update_payload": update_payload,
                    },
                )
                continue
            api.update_payout(resolved_payout.id, update_payload)
            matched_count += 1
        return found_count, matched_count

    def _read_secret(self, *, env_name: str, label: str) -> str:
        """Read a required secret from the environment."""
        value = os.environ.get(env_name)
        if value:
            return value
        msg = f"{label} was not provided. Set environment variable {env_name}."
        raise ValueError(msg)

    def _log_payout_update(
        self,
        payout_id: int | None,
        pay_date: date,
        confirm_payload: dict[str, Any],
        update_payload: dict[str, Any],
    ) -> None:
        """Print the worksheet-derived payloads for this payout."""
        print(
            f"Matched payout {payout_id or 'unconfirmed'} for {self._format_date(pay_date)} -> "
            f"confirm={json.dumps(None if confirm_payload is None else confirm_payload['payout'], sort_keys=True)} "
            f"update={json.dumps(update_payload['payout'], sort_keys=True)}",
        )

    def _build_confirm_payload(self, payout: PayoutRecord) -> dict[str, Any] | None:
        """Build the POST payload that confirms an unconfirmed payout."""
        if payout.holding_id is None or payout.company_event_id is None or payout.paid_on is None:
            return None
        return {
            "payout": {
                "holding_id": payout.holding_id,
                "company_event_id": payout.company_event_id,
                "paid_on": payout.paid_on.isoformat(),
                "state": self.confirmed_state,
            },
        }

    def _log_omitted_amount(self, payout_id: int | None, pay_date: date, amount: str) -> None:
        """Log when amount is omitted because the worksheet value is non-positive."""
        print(
            f"Omitting non-positive amount for payout {payout_id or 'unconfirmed'} "
            f"on {self._format_date(pay_date)}: amount={amount}",
        )

    def _read_worksheet_entries(self) -> tuple[dict[date, WorksheetEntry], list[dict[str, str]]]:
        """Load exact pay-date matches from the configured worksheet."""
        workbook = self.workbook_loader(self.excel_path, data_only=True, read_only=True)
        try:
            try:
                worksheet = workbook[self.worksheet_name]
            except KeyError as exc:
                raise ValueError(
                    f"Worksheet '{self.worksheet_name}' was not found in {self.excel_path}.",
                ) from exc

            header_row_index = self._locate_header_row(worksheet)
            entries: dict[date, WorksheetEntry] = {}
            skipped_rows: list[dict[str, str]] = []
            for row_index in range(header_row_index + 1, worksheet.max_row + 1):
                pay_date_value = worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Pay Date"]).value
                if pay_date_value in (None, ""):
                    continue

                try:
                    pay_date = self._normalize_excel_date(pay_date_value)
                except ValueError as exc:
                    skipped_rows.append(
                        {
                            "row": str(row_index),
                            "pay_date": self._describe_pay_date_value(pay_date_value),
                            "reason": str(exc),
                        },
                    )
                    continue

                if pay_date in entries:
                    msg = (
                        f"Worksheet '{self.worksheet_name}' contains duplicate Pay Date "
                        f"{self._format_date(pay_date)}."
                    )
                    raise ValueError(msg)

                try:
                    entries[pay_date] = WorksheetEntry(
                        pay_date=pay_date,
                        income_percent=self._normalize_decimal_string(
                            worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Inc%"]).value,
                        ),
                        income_amount=self._normalize_decimal_string(
                            worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Inc $"]).value,
                        ),
                        gross_amount=self._normalize_decimal_string(
                            worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Gross Amt"]).value,
                        ),
                        foreign_tax_withheld=self._normalize_decimal_string(
                            worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Tax"]).value,
                        ),
                        excess_tax=self._normalize_decimal_string(
                            worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Excess Tax"]).value,
                        ),
                        exchange_rate=self._normalize_decimal_string(
                            worksheet.cell(row=row_index, column=REQUIRED_HEADERS["FXRates"]).value,
                        ),
                    )
                except ValueError as exc:
                    skipped_rows.append(
                        {
                            "row": str(row_index),
                            "pay_date": self._format_date(pay_date),
                            "reason": str(exc),
                        },
                    )
        finally:
            workbook.close()

        return entries, skipped_rows

    def _determine_payout_date_window(
        self,
        worksheet_entries: dict[date, WorksheetEntry],
    ) -> tuple[date | None, date | None, bool]:
        """Return payout bounds and whether the API end date should be buffered.

        When bounds come from the worksheet, the returned end date is the latest
        Pay Date in Excel. Callers should pass that through
        `_sharesight_list_payouts_end_date` before `list_payouts`.
        """
        if self.payouts_start_date is not None or self.payouts_end_date is not None:
            return self.payouts_start_date, self.payouts_end_date, False
        if not worksheet_entries:
            return None, None, False
        pay_dates = sorted(worksheet_entries)
        return pay_dates[0], pay_dates[-1], True

    def _sharesight_list_payouts_end_date(
        self,
        end_date: date | None,
        *,
        buffer_worksheet_end: bool,
    ) -> date | None:
        """Extend worksheet-derived end dates so Sharesight includes the last pay day."""
        if end_date is None or not buffer_worksheet_end:
            return end_date
        return end_date + timedelta(days=PAYOUT_LIST_END_DATE_BUFFER_DAYS)

    def _locate_header_row(self, worksheet: Worksheet) -> int:
        """Find the worksheet row containing the required column headers."""
        expected_headers = set(REQUIRED_HEADERS)
        for row_index in range(1, worksheet.max_row + 1):
            values = {
                str(worksheet.cell(row=row_index, column=column_index).value).strip()
                for column_index in range(HEADER_START_COLUMN, max(REQUIRED_HEADERS.values()) + 1)
                if worksheet.cell(row=row_index, column=column_index).value is not None
                and str(worksheet.cell(row=row_index, column=column_index).value).strip()
            }
            if expected_headers.issubset(values):
                return row_index

        msg = (
            f"Worksheet '{self.worksheet_name}' is missing required headers: "
            f"{', '.join(REQUIRED_HEADERS)}."
        )
        raise ValueError(msg)

    def _normalize_excel_date(self, value: Any) -> date:
        """Normalize an Excel cell value into a calendar date."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, time):
            raise ValueError(f"Worksheet contains time-only value instead of date: {value}")
        parsed = datetime.strptime(str(value).strip(), DATE_OUTPUT_FORMAT)
        return parsed.date()

    def _coerce_optional_date(self, value: Any) -> date | None:
        """Coerce an optional configured date."""
        if value in (None, ""):
            return None
        if isinstance(value, date):
            return value
        for fmt in ("%Y-%m-%d", DATE_OUTPUT_FORMAT):
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Unable to parse configured date value: {value!r}")

    def _coerce_optional_bool(self, value: Any) -> bool | None:
        """Coerce an optional configured boolean value."""
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().casefold()
        if text in {"true", "1", "yes", "y", "on"}:
            return True
        if text in {"false", "0", "no", "n", "off"}:
            return False
        raise ValueError(f"Unable to parse configured boolean value: {value!r}")

    def _normalize_decimal_string(self, value: Any) -> str:
        """Normalize a numeric cell value into a Sharesight-friendly string."""
        if value is None or str(value).strip() == "":
            return "0"

        text_value = str(value).strip()
        if self._is_excel_error_value(text_value):
            raise ValueError(f"Worksheet contains Excel error value: {text_value}")

        try:
            numeric = Decimal(text_value)
        except InvalidOperation as exc:
            raise ValueError(f"Unable to parse numeric worksheet value: {value!r}") from exc

        text = format(numeric, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text or "0"

    def _format_date(self, value: date) -> str:
        """Format a date for summaries and Sharesight form entry."""
        return value.strftime(DATE_OUTPUT_FORMAT)

    def _describe_pay_date_value(self, value: Any) -> str:
        """Return a user-facing description of the raw worksheet pay-date cell."""
        if isinstance(value, datetime):
            return self._format_date(value.date())
        if isinstance(value, date):
            return self._format_date(value)
        return str(value).strip()

    def _is_excel_error_value(self, value: str) -> bool:
        """Return whether the supplied text looks like an Excel error literal."""
        return value in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NUM!", "#NAME?", "#NULL!"}

    def _to_float_or_none(self, value: Any) -> float | None:
        """Convert API numeric values to float for comparison."""
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _floats_differ(self, left: float | None, right: float | None, *, epsilon: float = 1e-9) -> bool:
        """Return whether two numeric values differ."""
        if left is None or right is None:
            return left != right
        return abs(left - right) > epsilon

    def _amount_should_be_included(self, amount: str) -> bool:
        """Return whether payout amount should be included in update payload."""
        try:
            return Decimal(str(amount).strip()) > 0
        except (InvalidOperation, ValueError):
            return False

class SharesightApiClient:
    """urllib-backed Sharesight API adapter."""

    def __init__(
        self,
        *,
        api_base_url: str,
        token_url: str,
        client_id: str,
        client_secret: str,
    ) -> None:
        self.api_base_url = api_base_url.rstrip("/")
        self.token_url = token_url
        self.client_id = client_id
        self.client_secret = client_secret
        self._access_token: str | None = None

    def resolve_portfolio(self, portfolio_name: str) -> PortfolioRecord:
        """Resolve the configured portfolio by exact or closest match."""
        response = self._request_json("GET", "/portfolios.json")
        portfolios = response.get("portfolios")
        if not isinstance(portfolios, list):
            raise RuntimeError("Unexpected Sharesight portfolios response: missing 'portfolios' list.")

        records = [
            PortfolioRecord(id=int(item["id"]), name=str(item["name"]))
            for item in portfolios
            if isinstance(item, dict) and item.get("id") is not None and item.get("name")
        ]
        for record in records:
            if record.name.casefold() == portfolio_name.casefold():
                return record

        closest_name = self._closest_portfolio_name(portfolio_name, [record.name for record in records])
        if closest_name is None:
            raise RuntimeError(
                f"Unable to resolve portfolio {portfolio_name!r}. "
                f"Visible portfolios: {[record.name for record in records]!r}",
            )
        for record in records:
            if record.name == closest_name:
                return record
        raise RuntimeError(f"Unable to resolve closest portfolio for {portfolio_name!r}.")

    def list_payouts(
        self,
        portfolio_id: int,
        *,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[PayoutRecord]:
        """Return payouts for the portfolio."""
        params: dict[str, str] = {}
        if start_date is not None:
            params["start_date"] = start_date.isoformat()
        if end_date is not None:
            params["end_date"] = end_date.isoformat()
        query = f"?{parse.urlencode(params)}" if params else ""
        response = self._request_json("GET", f"/portfolios/{portfolio_id}/payouts.json{query}")
        payouts = response.get("payouts")
        if not isinstance(payouts, list):
            raise RuntimeError("Unexpected Sharesight payouts response: missing 'payouts' list.")
        parsed_payouts: list[PayoutRecord] = []
        for item in payouts:
            if not isinstance(item, dict):
                continue
            parsed_payouts.append(self._parse_payout_record(item))
        return parsed_payouts

    def get_payout(self, payout_id: int) -> PayoutRecord:
        """Return one payout by ID."""
        response = self._request_json("GET", f"/payouts/{payout_id}.json")
        payout = response.get("payout", response)
        if not isinstance(payout, dict):
            raise RuntimeError(f"Unexpected payout response for payout {payout_id}.")
        parsed = self._parse_payout_record(payout)
        if parsed.id is None:
            raise RuntimeError(f"Unexpected payout response for payout {payout_id}: missing id.")
        return parsed

    def confirm_payout(self, payload: dict[str, Any]) -> PayoutRecord:
        """Confirm or create an unconfirmed payout."""
        response = self._request_json("POST", "/payouts.json", payload=payload)
        payout = response.get("payout", response)
        if not isinstance(payout, dict):
            raise RuntimeError("Unexpected response from Sharesight payout confirmation.")
        parsed = self._parse_payout_record(payout)
        if parsed.id is None:
            raise RuntimeError(f"Sharesight payout confirmation did not return a payout id: {response!r}")
        return parsed

    def update_payout(self, payout_id: int, payload: dict[str, Any]) -> dict[str, Any]:
        """Update one payout."""
        return self._request_json("PUT", f"/payouts/{payout_id}.json", payload=payload)

    def close(self) -> None:
        """Close API resources."""
        return None

    def _parse_payout_record(self, payload: dict[str, Any]) -> PayoutRecord:
        """Normalize one payout JSON object."""
        payout_id = payload.get("id")
        paid_on = payload.get("paid_on")
        return PayoutRecord(
            id=None if payout_id in (None, "") else int(payout_id),
            paid_on=None if paid_on in (None, "") else self._parse_api_date(str(paid_on)),
            state=str(payload.get("state", "")),
            raw=payload,
        )

    def _request_json(
        self,
        method: str,
        path_or_url: str,
        *,
        payload: dict[str, Any] | None = None,
        include_auth: bool = True,
        token_request: bool = False,
    ) -> dict[str, Any]:
        """Send an HTTP request and decode the JSON response."""
        url = path_or_url if path_or_url.startswith("http") else f"{self.api_base_url}{path_or_url}"
        data = None
        headers = {"Accept": "application/json"}
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
            headers["Content-Type"] = "application/json"
        if token_request:
            basic = b64encode(f"{self.client_id}:{self.client_secret}".encode("utf-8")).decode("ascii")
            headers["Authorization"] = f"Basic {basic}"
        elif include_auth:
            headers["Authorization"] = f"Bearer {self._get_access_token()}"

        req = request.Request(url, data=data, headers=headers, method=method)
        try:
            with request.urlopen(req) as response:
                body = response.read().decode("utf-8")
        except error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Sharesight API request failed: {method} {url} -> {exc.code}: {body}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"Unable to reach Sharesight API at {url}: {exc.reason}") from exc

        if body.strip() == "":
            return {}
        decoded = json.loads(body)
        if not isinstance(decoded, dict):
            raise RuntimeError(f"Unexpected non-object JSON response from Sharesight API: {decoded!r}")
        return decoded

    def _get_access_token(self) -> str:
        """Return an OAuth access token, fetching it on first use."""
        if self._access_token is not None:
            return self._access_token

        token_body = parse.urlencode({"grant_type": "client_credentials"}).encode("utf-8")
        response = self._request_token(token_body)
        access_token = response.get("access_token")
        if not access_token:
            raise RuntimeError("Sharesight OAuth token response did not include 'access_token'.")
        self._access_token = str(access_token)
        return self._access_token

    def _request_token(self, body: bytes) -> dict[str, Any]:
        """Fetch an OAuth token using client credentials."""
        basic = b64encode(f"{self.client_id}:{self.client_secret}".encode("utf-8")).decode("ascii")
        headers = {
            "Accept": "application/json",
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
        }
        req = request.Request(self.token_url, data=body, headers=headers, method="POST")
        try:
            with request.urlopen(req) as response:
                raw = response.read().decode("utf-8")
        except error.HTTPError as exc:
            body_text = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"Sharesight OAuth token request failed: POST {self.token_url} -> {exc.code}: {body_text}",
            ) from exc
        except error.URLError as exc:
            raise RuntimeError(f"Unable to reach Sharesight OAuth endpoint at {self.token_url}: {exc.reason}") from exc

        decoded = json.loads(raw)
        if not isinstance(decoded, dict):
            raise RuntimeError(f"Unexpected token response from Sharesight OAuth endpoint: {decoded!r}")
        return decoded

    def _closest_portfolio_name(self, target: str, candidates: list[str]) -> str | None:
        """Return the closest visible portfolio name above a similarity threshold."""
        best_name: str | None = None
        best_score = 0.0
        for candidate in candidates:
            score = SequenceMatcher(None, target.casefold(), candidate.casefold()).ratio()
            if score > best_score:
                best_score = score
                best_name = candidate
        if best_score >= 0.6:
            return best_name
        return None

    def _parse_api_date(self, value: str) -> date:
        """Parse an API date string."""
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
