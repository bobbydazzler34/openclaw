"""Unit tests for the MSTY tracker skill."""

from __future__ import annotations

from datetime import date
from pathlib import Path
import tempfile
import unittest
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pandas as pd
import requests

from openclaw.skills.msty_tracker.skill import DistributionRecord, MstyTrackerSkill


PUBLIC_TEST_WORKBOOK_PATH = Path("/Users/chriscropley/Public/Personal CashFlow.xlsx")


def build_website_table() -> pd.DataFrame:
    """Return a minimal distribution table shaped like the website."""
    return pd.DataFrame(
        [
            {
                "DISTRIBUTION PER SHARE": "$0.4091",
                "DECLARED DATE": "12/31/2025",
                "EX DATE": "01/02/2026",
                "RECORD DATE": "01/02/2026",
                "PAYABLE DATE": "01/05/2026",
                "ROC": "94.34%",
            },
            {
                "DISTRIBUTION PER SHARE": "$0.5106",
                "DECLARED DATE": "12/24/2025",
                "EX DATE": "12/26/2025",
                "RECORD DATE": "12/26/2025",
                "PAYABLE DATE": "12/29/2025",
                "ROC": "80.04%",
            },
            {
                "DISTRIBUTION PER SHARE": "$0.1352",
                "DECLARED DATE": "11/26/2025",
                "EX DATE": "11/28/2025",
                "RECORD DATE": "11/28/2025",
                "PAYABLE DATE": "12/01/2025",
                "ROC": "69.67%",
            },
            {
                "DISTRIBUTION PER SHARE": "$0.5859",
                "DECLARED DATE": "12/10/2025",
                "EX DATE": "12/11/2025",
                "RECORD DATE": "12/11/2025",
                "PAYABLE DATE": "12/12/2025",
                "ROC": "96.35%",
            },
            {
                "DISTRIBUTION PER SHARE": "$0.9999",
                "DECLARED DATE": "12/17/2025",
                "EX DATE": "12/18/2025",
                "RECORD DATE": "12/18/2025",
                "PAYABLE DATE": "12/19/2025",
                "ROC": "0.00%",
            },
        ],
    )


def build_worksheet_table() -> pd.DataFrame:
    """Return a raw worksheet layout with the real header row below top padding."""
    return pd.DataFrame(
        [
            [None, "DISTRIBUTIONS", None, None, None, None],
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
            [
                None,
                "Distros",
                "declared date",
                "ex date",
                "record date",
                "payable date",
            ],
            [None, 0.1352, date(2025, 11, 26), date(2025, 11, 27), date(2025, 11, 27), date(2025, 11, 28)],
            [None, 0.5859, date(2025, 12, 10), date(2025, 12, 11), date(2025, 12, 11), date(2025, 12, 12)],
            [None, 0.5106, date(2025, 12, 24), date(2025, 12, 25), date(2025, 12, 25), date(2025, 12, 26)],
            [None, 0.4091, date(2025, 12, 31), date(2026, 1, 1), date(2026, 1, 1), date(2026, 1, 2)],
        ],
        columns=list(range(6)),
    )


def create_test_workbook(workbook_path: Path) -> None:
    """Create a workbook fixture matching the raw worksheet layout."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Distributions"

    rows = build_worksheet_table().values.tolist()
    for row_index, values in enumerate(rows, start=1):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column_index).value = value

    for row_index in range(5, 9):
        for column_index in range(2, 7):
            worksheet.cell(row=row_index, column=column_index).font = Font(bold=(row_index == 8))
            worksheet.cell(row=row_index, column=column_index).number_format = "dd/mmm/yyyy" if column_index >= 3 else "0.0000"

    dc_sheet = workbook.create_sheet("CS FY2526")
    dc_sheet.cell(row=6, column=8).value = "MSTY Distributions USD 2025/2026"
    headers = [
        "Pay Date",
        "Dist / Share",
        "ROC%",
        "Inc%",
        "ROC $",
        "Inc $",
        "Gross Amt",
        "Tax",
        "GA-Tax",
        "Excess Tax",
        "Act Tax",
        "FXRates",
    ]
    for offset, header in enumerate(headers, start=8):
        dc_sheet.cell(row=7, column=offset).value = header
    dc_sheet["H8"] = "=Distributions!F5"
    dc_sheet["I8"] = "=Distributions!B5"
    dc_sheet["J8"] = None
    dc_sheet["K8"] = "=100-J8"
    dc_sheet["L8"] = "=J8*1"
    dc_sheet["M8"] = "=K8*1"
    dc_sheet["N8"] = "=SUM(L8:M8)"
    dc_sheet["O8"] = "=N8"
    dc_sheet["P8"] = "=N8"
    dc_sheet["Q8"] = "=L8*0"
    dc_sheet["R8"] = "=M8*0"
    dc_sheet["S8"] = "=1"
    for column_index in range(8, 20):
        dc_sheet.cell(row=8, column=column_index).font = Font(bold=False)
        dc_sheet.cell(row=8, column=column_index).number_format = "0.00" if column_index == 10 else "General"

    workbook.save(workbook_path)
    workbook.close()


def create_final_state_workbook(workbook_path: Path) -> None:
    """Create a workbook fixture with final MSTY rows plus DC Pavula formulas."""
    workbook = Workbook()
    distributions = workbook.active
    distributions.title = "Distributions"

    distribution_rows = [
        [None, "DISTRIBUTIONS", None, None, None, None],
        [None, None, None, None, None, None],
        [None, None, None, None, None, None],
        [None, "Distros", "declared date", "ex date", "record date", "payable date"],
        [None, 0.1352, date(2025, 11, 26), date(2025, 11, 28), date(2025, 11, 28), date(2025, 12, 1)],
        [None, 0.5859, date(2025, 12, 10), date(2025, 12, 11), date(2025, 12, 11), date(2025, 12, 12)],
        [None, 0.9999, date(2025, 12, 17), date(2025, 12, 18), date(2025, 12, 18), date(2025, 12, 19)],
        [None, 0.5106, date(2025, 12, 24), date(2025, 12, 26), date(2025, 12, 26), date(2025, 12, 29)],
        [None, 0.4091, date(2025, 12, 31), date(2026, 1, 2), date(2026, 1, 2), date(2026, 1, 5)],
    ]
    for row_index, values in enumerate(distribution_rows, start=1):
        for column_index, value in enumerate(values, start=1):
            distributions.cell(row=row_index, column=column_index).value = value

    dc_sheet = workbook.create_sheet("CS FY2526")
    dc_sheet.cell(row=6, column=8).value = "MSTY Distributions USD 2025/2026"
    headers = [
        "Pay Date",
        "Dist / Share",
        "ROC%",
        "Inc%",
        "ROC $",
        "Inc $",
        "Gross Amt",
        "Tax",
        "GA-Tax",
        "Excess Tax",
        "Act Tax",
        "FXRates",
    ]
    for offset, header in enumerate(headers, start=8):
        dc_sheet.cell(row=7, column=offset).value = header

    data_rows = {
        8: (5, None),
        9: (6, 96.35),
        10: (8, None),
        11: (9, None),
    }
    for row_index, (distribution_row, roc_value) in data_rows.items():
        dc_sheet.cell(row=row_index, column=8).value = f"=Distributions!F{distribution_row}"
        dc_sheet.cell(row=row_index, column=9).value = f"=Distributions!B{distribution_row}"
        dc_sheet.cell(row=row_index, column=10).value = roc_value
        dc_sheet.cell(row=row_index, column=11).value = f"=100-J{row_index}"
        dc_sheet.cell(row=row_index, column=12).value = f"=J{row_index}*1"
        dc_sheet.cell(row=row_index, column=13).value = f"=K{row_index}*1"
        dc_sheet.cell(row=row_index, column=14).value = f"=SUM(L{row_index}:M{row_index})"
        dc_sheet.cell(row=row_index, column=15).value = f"=N{row_index}-0"
        dc_sheet.cell(row=row_index, column=16).value = f"=N{row_index}-0"
        dc_sheet.cell(row=row_index, column=17).value = f"=L{row_index}*0"
        dc_sheet.cell(row=row_index, column=18).value = f"=M{row_index}*0"
        dc_sheet.cell(row=row_index, column=19).value = f"=1"

    for row_index in range(8, 12):
        for column_index in range(8, 20):
            dc_sheet.cell(row=row_index, column=column_index).font = Font(bold=False)
            if column_index == 10:
                dc_sheet.cell(row=row_index, column=column_index).number_format = "0.00"
            else:
                dc_sheet.cell(row=row_index, column=column_index).number_format = "General"

    workbook.save(workbook_path)
    workbook.close()


class TestMstyTrackerSkill(unittest.TestCase):
    """Unit coverage for the MSTY tracker skill."""

    def test_run_returns_missing_and_different_distributions(self) -> None:
        """Website rows are bucketed by missing Distros vs changed declared/ex dates."""
        response = Mock()
        response.text = "<html>ignored</html>"
        response.raise_for_status.return_value = None

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_test_workbook(workbook_path)

            with (
                patch("openclaw.skills.msty_tracker.skill.requests.get", return_value=response),
                patch(
                    "openclaw.skills.msty_tracker.skill.pd.read_html",
                    return_value=[build_website_table()],
                ),
            ):
                skill = MstyTrackerSkill(excel_path=workbook_path)
                result = skill.run()

            updated_workbook = load_workbook(workbook_path)
            updated_sheet = updated_workbook["Distributions"]
            inserted_row = [updated_sheet.cell(row=7, column=column).value for column in range(2, 7)]
            updated_1352_row = [updated_sheet.cell(row=5, column=column).value for column in range(2, 7)]
            updated_5106_row = [updated_sheet.cell(row=8, column=column).value for column in range(2, 7)]
            updated_4091_row = [updated_sheet.cell(row=9, column=column).value for column in range(2, 7)]
            inserted_date_format = updated_sheet.cell(row=7, column=3).number_format
            inserted_distro_format = updated_sheet.cell(row=7, column=2).number_format
            updated_workbook.close()

        self.assertEqual(result["status"], "success")
        self.assertIn("dc_skipped_no_dc_row", result)
        self.assertIn("dc_merge_warnings", result)
        self.assertEqual(result["dc_merge_warnings"], [])
        self.assertEqual(result["rows_found"], 5)
        self.assertEqual(result["rows_missing"], 1)
        self.assertEqual(result["rows_different"], 3)
        self.assertEqual(result["rows_matching"], 1)
        self.assertEqual(result["rows_inserted"], 1)
        self.assertEqual(result["rows_updated"], 3)
        self.assertListEqual(
            result["missing_distributions"],
            [
                {
                    "Distros": "0.9999",
                    "declared date": "17/Dec/2025",
                    "ex date": "18/Dec/2025",
                    "record date": "18/Dec/2025",
                    "payable date": "19/Dec/2025",
                },
            ],
        )
        self.assertEqual(inserted_row[0], 0.9999)
        self.assertEqual(inserted_row[1].date(), date(2025, 12, 17))
        self.assertEqual(inserted_row[2].date(), date(2025, 12, 18))
        self.assertEqual(inserted_row[3].date(), date(2025, 12, 18))
        self.assertEqual(inserted_row[4].date(), date(2025, 12, 19))
        self.assertEqual(inserted_date_format, "dd/mmm/yyyy")
        self.assertEqual(inserted_distro_format, "0.0000")
        self.assertEqual(updated_1352_row[2].date(), date(2025, 11, 28))
        self.assertEqual(updated_5106_row[2].date(), date(2025, 12, 26))
        self.assertEqual(updated_4091_row[2].date(), date(2026, 1, 2))
        self.assertListEqual(
            result["different_distributions"],
            [
                {
                    "Distros": "0.4091",
                    "declared date": "31/Dec/2025",
                    "ex date": "02/Jan/2026",
                    "record date": "02/Jan/2026",
                    "payable date": "05/Jan/2026",
                },
                {
                    "Distros": "0.5106",
                    "declared date": "24/Dec/2025",
                    "ex date": "26/Dec/2025",
                    "record date": "26/Dec/2025",
                    "payable date": "29/Dec/2025",
                },
                {
                    "Distros": "0.1352",
                    "declared date": "26/Nov/2025",
                    "ex date": "28/Nov/2025",
                    "record date": "28/Nov/2025",
                    "payable date": "01/Dec/2025",
                },
            ],
        )
        self.assertListEqual(
            result["matching_distributions"],
            [
                {
                    "Distros": "0.5859",
                    "declared date": "10/Dec/2025",
                    "ex date": "11/Dec/2025",
                    "record date": "11/Dec/2025",
                    "payable date": "12/Dec/2025",
                },
            ],
        )

    def test_run_writes_obsidian_log_when_enabled(self) -> None:
        """A markdown run log is emitted when Obsidian logging is enabled."""
        response = Mock()
        response.text = "<html>ignored</html>"
        response.raise_for_status.return_value = None

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            logs_dir = Path(temp_dir) / "logs"
            create_test_workbook(workbook_path)

            with (
                patch("openclaw.skills.msty_tracker.skill.requests.get", return_value=response),
                patch(
                    "openclaw.skills.msty_tracker.skill.pd.read_html",
                    return_value=[build_website_table()],
                ),
            ):
                skill = MstyTrackerSkill(excel_path=workbook_path)
                skill.config["obsidian_log_enabled"] = True
                skill.config["obsidian_log_dir"] = str(logs_dir)
                skill.config["obsidian_log_user"] = "bobbyd"
                skill.config["obsidian_log_operator"] = "Chris Cropley"
                skill.config["obsidian_log_environment"] = "test"
                skill.run()

            log_files = list(logs_dir.glob("*.md"))
            self.assertEqual(len(log_files), 1)
            self.assertRegex(log_files[0].name, r"\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2}Z_msty_tracker_bobbyd\.md")
            log_content = log_files[0].read_text(encoding="utf-8")
            self.assertIn("skill_id: msty_tracker", log_content)
            self.assertIn('status: "success"', log_content)

    def test_run_raises_runtime_error_for_network_failure(self) -> None:
        """Network failures are surfaced as runtime errors."""
        with patch(
            "openclaw.skills.msty_tracker.skill.requests.get",
            side_effect=requests.RequestException("boom"),
        ):
            skill = MstyTrackerSkill()
            with self.assertRaisesRegex(RuntimeError, "Unable to download MSTY distribution data."):
                skill.run()

    def test_write_dc_pavula_sheet_fills_missing_and_blank_roc_rows(self) -> None:
        """DC Pavula rows are inserted or updated from website ROC values."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_final_state_workbook(workbook_path)
            skill = MstyTrackerSkill(
                excel_path=workbook_path,
                dc_pavula_insert_missing_rows=True,
            )

            rows_inserted, rows_updated, pending_entries, skipped, merge_warnings = (
                skill._write_dc_pavula_sheet(
                [
                    DistributionRecord(
                        distros="0.1352",
                        declared_date="26/Nov/2025",
                        ex_date="28/Nov/2025",
                        record_date="28/Nov/2025",
                        payable_date="01/Dec/2025",
                        roc_percent=69.67,
                    ),
                    DistributionRecord(
                        distros="0.5859",
                        declared_date="10/Dec/2025",
                        ex_date="11/Dec/2025",
                        record_date="11/Dec/2025",
                        payable_date="12/Dec/2025",
                        roc_percent=96.35,
                    ),
                    DistributionRecord(
                        distros="0.9999",
                        declared_date="17/Dec/2025",
                        ex_date="18/Dec/2025",
                        record_date="18/Dec/2025",
                        payable_date="19/Dec/2025",
                        roc_percent=0.0,
                    ),
                    DistributionRecord(
                        distros="0.5106",
                        declared_date="24/Dec/2025",
                        ex_date="26/Dec/2025",
                        record_date="26/Dec/2025",
                        payable_date="29/Dec/2025",
                        roc_percent=80.04,
                    ),
                    DistributionRecord(
                        distros="0.4091",
                        declared_date="31/Dec/2025",
                        ex_date="02/Jan/2026",
                        record_date="02/Jan/2026",
                        payable_date="05/Jan/2026",
                        roc_percent=94.34,
                    ),
                ],
                )
            )

            self.assertEqual(skipped, [])
            self.assertEqual(merge_warnings, [])

            workbook = load_workbook(workbook_path)
            dc_sheet = workbook["CS FY2526"]
            workbook.close()

        self.assertEqual(rows_inserted, 1)
        self.assertEqual(rows_updated, 4)
        self.assertEqual(
            pending_entries,
            [
                {"Pay Date": "01/Dec/2025", "ROC%": 69.67},
                {"Pay Date": "12/Dec/2025", "ROC%": 96.35},
                {"Pay Date": "19/Dec/2025", "ROC%": 0.0},
                {"Pay Date": "29/Dec/2025", "ROC%": 80.04},
                {"Pay Date": "05/Jan/2026", "ROC%": 94.34},
            ],
        )
        self.assertEqual(dc_sheet["J8"].value, 69.67)
        self.assertEqual(dc_sheet["J9"].value, 96.35)
        self.assertEqual(dc_sheet["H10"].value, "=Distributions!F7")
        self.assertEqual(dc_sheet["I10"].value, "=Distributions!B7")
        self.assertEqual(dc_sheet["J10"].value, 0.0)
        self.assertEqual(dc_sheet["J11"].value, 80.04)
        self.assertEqual(dc_sheet["J12"].value, 94.34)

    def test_write_dc_pavula_merge_preflight_reports_overlap(self) -> None:
        """Merged ranges spanning H–T on a DC row produce dc_merge_warnings."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_test_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            workbook["CS FY2526"].merge_cells("H8:I8")
            workbook.save(workbook_path)
            workbook.close()

            skill = MstyTrackerSkill(excel_path=workbook_path)
            _inserted, _updated, _pending, _skipped, merge_warnings = skill._write_dc_pavula_sheet(
                [
                    DistributionRecord(
                        distros="0.1352",
                        declared_date="26/Nov/2025",
                        ex_date="27/Nov/2025",
                        record_date="27/Nov/2025",
                        payable_date="28/Nov/2025",
                        roc_percent=69.67,
                    ),
                ],
            )

        self.assertEqual(len(merge_warnings), 1)
        self.assertEqual(merge_warnings[0]["row"], 8)
        self.assertEqual(merge_warnings[0]["dc_columns"], "H:T")
        self.assertIn("H8", merge_warnings[0]["merged_range"])
        self.assertIn("I8", merge_warnings[0]["merged_range"])

    def test_run_raises_clear_error_for_missing_worksheet(self) -> None:
        """A missing worksheet raises a clear error."""
        response = Mock()
        response.text = "<html>ignored</html>"
        response.raise_for_status.return_value = None

        with (
            patch("openclaw.skills.msty_tracker.skill.requests.get", return_value=response),
            patch("openclaw.skills.msty_tracker.skill.pd.read_html", return_value=[build_website_table()]),
            patch(
                "openclaw.skills.msty_tracker.skill.pd.read_excel",
                side_effect=ValueError("Worksheet named 'Distributions' not found"),
            ),
        ):
            skill = MstyTrackerSkill(worksheet_name="Distributions")
            with self.assertRaisesRegex(ValueError, "Worksheet 'Distributions' was not found"):
                skill.run()

    def test_run_raises_clear_error_for_missing_columns(self) -> None:
        """A worksheet missing required columns raises a clear error."""
        response = Mock()
        response.text = "<html>ignored</html>"
        response.raise_for_status.return_value = None

        worksheet = build_worksheet_table()
        worksheet.iloc[3, 4] = "wrong record header"
        with (
            patch("openclaw.skills.msty_tracker.skill.requests.get", return_value=response),
            patch("openclaw.skills.msty_tracker.skill.pd.read_html", return_value=[build_website_table()]),
            patch("openclaw.skills.msty_tracker.skill.pd.read_excel", return_value=worksheet),
        ):
            skill = MstyTrackerSkill()
            with self.assertRaisesRegex(ValueError, "missing required columns"):
                skill.run()

    @unittest.skipUnless(PUBLIC_TEST_WORKBOOK_PATH.exists(), "Public workbook not present.")
    def test_public_workbook_check(self) -> None:
        """Integration-style check against a workbook in /Users/chriscropley/Public."""
        try:
            import openpyxl  # noqa: F401
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed; pandas can't read xlsx.")

        response = Mock()
        response.text = "<html>ignored</html>"
        response.raise_for_status.return_value = None

        with (
            patch("openclaw.skills.msty_tracker.skill.requests.get", return_value=response),
            patch("openclaw.skills.msty_tracker.skill.pd.read_html", return_value=[build_website_table()]),
        ):
            skill = MstyTrackerSkill(excel_path=PUBLIC_TEST_WORKBOOK_PATH)
            try:
                result = skill.run()
            except ValueError as exc:
                self.skipTest(f"Public workbook isn't compatible: {exc}")
            except PermissionError as exc:
                self.skipTest(f"Public workbook isn't writable in test sandbox: {exc}")

        self.assertEqual(result["status"], "success")
        self.assertIn("missing_distributions", result)


if __name__ == "__main__":
    unittest.main()
