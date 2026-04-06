"""MSTY tracker skill implementation."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from io import StringIO
import os
from pathlib import Path
import re
import shutil
from typing import Any, Callable
from pprint import pformat

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import requests

from openclaw.skills._base.skill_base import SkillBase

MSTY_URL = "https://yieldmaxetfs.com/our-etfs/msty/"
EXCEL_PATH = "~/OneDrive/Documents/Finance/Personal CashFlow.xlsx"
EXCEL_PATH_ENV_VAR = "OPENCLAW_MSTY_TRACKER_EXCEL_PATH"
WORKSHEET_NAME = "Distributions"
DC_PAVULA_WORKSHEET_NAME = "CS FY2526"
# openpyxl insert_rows + row copy breaks many real workbooks (tables, array formulas, merges).
# Safe default: only fill ROC% on rows that already reference Distributions.
DEFAULT_UPDATE_DC_PAVULA = True
DEFAULT_DC_PAVULA_INSERT_MISSING_ROWS = False
REQUEST_TIMEOUT_SECONDS = 30
REQUEST_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/136.0.0.0 Safari/537.36"
    ),
}

WORKSHEET_COLUMNS = [
    "Distros",
    "declared date",
    "ex date",
    "record date",
    "payable date",
]
SOURCE_COLUMNS = [
    "DISTRIBUTION PER SHARE",
    "DECLARED DATE",
    "EX DATE",
    "RECORD DATE",
    "PAYABLE DATE",
    "ROC",
]
DATE_OUTPUT_FORMAT = "%d/%b/%Y"
DATE_PARSE_FORMAT = "%d/%b/%Y"
DC_PAVULA_HEADERS = [
    "Pay Date",
    "Dist / Share",
    "ROC%",
    "Inc%",
    "ROC $",
    "Inc $",
    "Gross Amt",
    "Tax",
    "GA-Tax",
    "Excess Tax",
    "Act Tax",
    "FXRates",
]
DC_PAVULA_START_COLUMN = 8
DC_PAVULA_END_COLUMN = 20
DC_PAVULA_ROC_COLUMN = 10
DIST_SHEET_REF_PATTERN = re.compile(r"Distributions!\$?[BF]\$?(\d+)")


@dataclass(slots=True)
class DistributionRecord:
    """Normalized worksheet-shaped distribution row."""

    distros: str
    declared_date: str
    ex_date: str
    record_date: str
    payable_date: str
    roc_percent: float | None = None

    def as_dict(self) -> dict[str, str]:
        """Return a worksheet-compatible mapping."""
        return {
            "Distros": self.distros,
            "declared date": self.declared_date,
            "ex date": self.ex_date,
            "record date": self.record_date,
            "payable date": self.payable_date,
        }

    def date_key(self) -> tuple[str, str, str, str]:
        """Return the tuple used to identify an existing worksheet row."""
        return (
            self.declared_date,
            self.ex_date,
            self.record_date,
            self.payable_date,
        )

    def declared_ex_key(self) -> tuple[str, str]:
        """Return the date fields used to detect changed distributions."""
        return (self.declared_date, self.ex_date)

    def distribution_key(self) -> tuple[str, str, str, str, str]:
        """Return the full identifying key for the live Distributions sheet row."""
        return (
            self.distros,
            self.declared_date,
            self.ex_date,
            self.record_date,
            self.payable_date,
        )

    def dc_entry(self) -> dict[str, str | float]:
        """Return the DC Pavula write payload for logging/output."""
        return {
            "Pay Date": self.payable_date,
            "ROC%": 0.0 if self.roc_percent is None else self.roc_percent,
        }


class MstyTrackerSkill(SkillBase):
    """Fetch MSTY distributions and report worksheet rows that are still missing."""

    def __init__(
        self,
        config_path: str | Path | None = None,
        *,
        source_url: str | None = None,
        excel_path: str | Path | None = None,
        worksheet_name: str | None = None,
        requests_get: Callable[..., requests.Response] | None = None,
        read_html: Callable[..., list[pd.DataFrame]] | None = None,
        read_excel: Callable[..., pd.DataFrame] | None = None,
        workbook_loader: Callable[..., Any] | None = None,
        update_dc_pavula: bool | None = None,
        dc_pavula_insert_missing_rows: bool | None = None,
    ) -> None:
        """Initialize the skill with optional dependency injection for testing."""
        super().__init__(config_path)
        self.source_url = str(self.config.get("source_url", source_url or MSTY_URL))
        excel_path_env_var = str(
            self.config.get("excel_path_env_var", EXCEL_PATH_ENV_VAR),
        )
        configured_excel_path = (
            excel_path
            or os.getenv(excel_path_env_var)
            or self.config.get("excel_path", EXCEL_PATH)
        )
        self.excel_path = self._resolve_excel_path(configured_excel_path)
        self.worksheet_name = str(
            worksheet_name or self.config.get("worksheet_name", WORKSHEET_NAME),
        )
        self.requests_get = requests_get or requests.get
        self.read_html = read_html or pd.read_html
        self.read_excel = read_excel or pd.read_excel
        self.workbook_loader = workbook_loader or load_workbook
        self.update_dc_pavula = (
            bool(update_dc_pavula)
            if update_dc_pavula is not None
            else bool(self.config.get("update_dc_pavula", DEFAULT_UPDATE_DC_PAVULA))
        )
        self.dc_pavula_insert_missing_rows = (
            bool(dc_pavula_insert_missing_rows)
            if dc_pavula_insert_missing_rows is not None
            else bool(
                self.config.get(
                    "dc_pavula_insert_missing_rows",
                    DEFAULT_DC_PAVULA_INSERT_MISSING_ROWS,
                ),
            )
        )
        # Set when we make a pre-write backup for DC Pavula.
        self._dc_backup_path: Path | None = None

    def _backup_excel_for_dc_pavula(self) -> Path:
        """Copy the workbook before DC Pavula writes (for rollback safety)."""
        source_path = self.excel_path
        if not source_path.exists():
            raise FileNotFoundError(f"Workbook not found: {source_path}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = f"{source_path.stem}_backup_dc_pavula_{timestamp}{source_path.suffix}"
        backup_path = source_path.with_name(base)

        # Avoid overwriting if multiple runs happen within the same second.
        attempt = 1
        while backup_path.exists():
            backup_path = source_path.with_name(
                f"{source_path.stem}_backup_dc_pavula_{timestamp}_{attempt}{source_path.suffix}",
            )
            attempt += 1

        shutil.copy2(source_path, backup_path)
        return backup_path

    def _resolve_excel_path(self, configured_path: str | Path) -> Path:
        """Resolve the workbook path across common OneDrive folder naming patterns."""
        candidate = Path(configured_path).expanduser()
        if candidate.exists():
            return candidate

        candidate_parts = candidate.parts
        home_dir = Path.home()
        one_drive_segment = next(
            (part for part in candidate_parts if part == "OneDrive" or part.startswith("OneDrive - ")),
            None,
        )
        if one_drive_segment is None:
            return candidate

        relative_parts = []
        found_segment = False
        for part in candidate_parts:
            if found_segment:
                relative_parts.append(part)
            elif part == one_drive_segment:
                found_segment = True

        one_drive_roots = sorted(
            path
            for path in home_dir.iterdir()
            if path.is_dir() and (path.name == "OneDrive" or path.name.startswith("OneDrive - "))
        )
        for one_drive_root in one_drive_roots:
            resolved = one_drive_root.joinpath(*relative_parts)
            if resolved.exists():
                return resolved

        return candidate

    def run(self) -> dict[str, Any]:
        """Fetch MSTY data and return missing worksheet rows without editing the workbook."""
        self.logger.info("Starting MSTY distribution comparison.")
        source_html = self._download_source_html()
        website_records = self._read_website_distributions(source_html)
        worksheet_records = self._read_existing_worksheet()
        missing_records, different_records, matching_records = self._partition_records(
            website_records,
            worksheet_records,
        )
        rows_inserted, rows_updated = self._write_sheet(missing_records, different_records)
        self._dc_backup_path = None
        if self.update_dc_pavula:
            (
                dc_rows_inserted,
                dc_rows_updated,
                dc_pending_entries,
                dc_skipped_no_dc_row,
                dc_merge_warnings,
            ) = self._write_dc_pavula_sheet(website_records)
        else:
            dc_rows_inserted = 0
            dc_rows_updated = 0
            dc_pending_entries = []
            dc_skipped_no_dc_row = []
            dc_merge_warnings = []
            self._dc_backup_path = None
        print("Matching distributions (already in worksheet):")
        print(pformat([record.as_dict() for record in matching_records]))
        print("Different distributions (same Distros, changed declared/ex date):")
        print(pformat([record.as_dict() for record in different_records]))
        print("DC Pavula rows to enter (Pay Date, ROC%):")
        print(pformat(dc_pending_entries))
        if dc_skipped_no_dc_row:
            print(
                "DC Pavula skipped (no existing row linking to that Distributions row; "
                "add a row in Excel or set dc_pavula_insert_missing_rows: true):",
            )
            print(pformat(dc_skipped_no_dc_row))
        if dc_merge_warnings:
            print("DC Pavula pre-flight: merged cells overlap columns H–T on rows we are writing:")
            print(pformat(dc_merge_warnings))
        self.logger.info(
            "MSTY distribution comparison completed with %s missing rows and %s different rows.",
            len(missing_records),
            len(different_records),
        )
        return {
            "status": "success",
            "source_url": self.source_url,
            "workbook_path": str(self.excel_path),
            "worksheet_name": self.worksheet_name,
            "rows_found": len(website_records),
            "rows_missing": len(missing_records),
            "rows_different": len(different_records),
            "rows_matching": len(matching_records),
            "rows_inserted": rows_inserted,
            "rows_updated": rows_updated,
            "dc_rows_inserted": dc_rows_inserted,
            "dc_rows_updated": dc_rows_updated,
            "dc_skipped_no_dc_row": dc_skipped_no_dc_row,
            "dc_merge_warnings": dc_merge_warnings,
            "dc_backup_path": str(self._dc_backup_path) if self._dc_backup_path else None,
            "missing_distributions": [record.as_dict() for record in missing_records],
            "different_distributions": [record.as_dict() for record in different_records],
            "matching_distributions": [record.as_dict() for record in matching_records],
            "dc_pending_entries": dc_pending_entries,
        }

    def _download_source_html(self) -> str:
        """Download the MSTY ETF page."""
        self.logger.info("Downloading MSTY page from %s.", self.source_url)
        try:
            response = self.requests_get(
                self.source_url,
                timeout=REQUEST_TIMEOUT_SECONDS,
                headers=REQUEST_HEADERS,
            )
            response.raise_for_status()
        except requests.RequestException as exc:
            self.logger.exception("Failed to download MSTY page.")
            raise RuntimeError("Unable to download MSTY distribution data.") from exc
        return response.text

    def _read_website_distributions(self, html: str) -> list[DistributionRecord]:
        """Parse the website distributions table into normalized worksheet rows."""
        self.logger.info("Parsing MSTY distributions table from downloaded HTML.")
        try:
            tables = self.read_html(StringIO(html))
        except ValueError as exc:
            self.logger.exception("No HTML tables found on MSTY page.")
            raise ValueError("Unable to locate a distributions table on the MSTY page.") from exc

        distribution_table = self._select_distribution_table(tables)
        normalized_table = distribution_table.loc[:, SOURCE_COLUMNS].copy()
        records = [
            DistributionRecord(
                distros=self._normalize_distribution_value(row["DISTRIBUTION PER SHARE"]),
                declared_date=self._normalize_date(row["DECLARED DATE"]),
                ex_date=self._normalize_date(row["EX DATE"]),
                record_date=self._normalize_date(row["RECORD DATE"]),
                payable_date=self._normalize_date(row["PAYABLE DATE"]),
                roc_percent=self._normalize_roc_percent(row["ROC"]),
            )
            for _, row in normalized_table.iterrows()
        ]
        self.logger.info("Parsed %s distribution rows from the website.", len(records))
        return records

    def _select_distribution_table(self, tables: list[pd.DataFrame]) -> pd.DataFrame:
        """Return the website table that matches the expected distribution columns."""
        expected = set(SOURCE_COLUMNS)
        for table in tables:
            normalized_columns = {str(column).strip() for column in table.columns}
            if expected.issubset(normalized_columns):
                return table
        msg = "Unable to find the MSTY distributions table in the website HTML."
        raise ValueError(msg)

    def _read_existing_worksheet(self) -> list[DistributionRecord]:
        """Load and normalize the existing worksheet rows."""
        self.logger.info(
            "Reading existing distribution data from %s [%s].",
            self.excel_path,
            self.worksheet_name,
        )
        try:
            raw_worksheet = self.read_excel(
                self.excel_path,
                sheet_name=self.worksheet_name,
                header=None,
            )
        except ValueError as exc:
            self.logger.exception("Configured worksheet is missing.")
            raise ValueError(
                f"Worksheet '{self.worksheet_name}' was not found in {self.excel_path}.",
            ) from exc

        worksheet = self._extract_distribution_table(raw_worksheet)
        normalized_rows = worksheet.loc[:, WORKSHEET_COLUMNS].copy().dropna(how="all")
        records: list[DistributionRecord] = []
        for _, row in normalized_rows.iterrows():
            if self._row_has_blank_dates(row):
                continue
            records.append(
                DistributionRecord(
                    distros=self._normalize_distribution_value(row["Distros"]),
                    declared_date=self._normalize_excel_date(row["declared date"]),
                    ex_date=self._normalize_excel_date(row["ex date"]),
                    record_date=self._normalize_excel_date(row["record date"]),
                    payable_date=self._normalize_excel_date(row["payable date"]),
                ),
            )
        self.logger.info("Loaded %s comparable worksheet rows.", len(records))
        return records

    def _extract_distribution_table(self, raw_worksheet: pd.DataFrame) -> pd.DataFrame:
        """Detect the header row and return a normalized worksheet table."""
        header_row_index = self._find_header_row_index(raw_worksheet)
        header_values = [
            str(value).strip() if not pd.isna(value) else ""
            for value in raw_worksheet.iloc[header_row_index].tolist()
        ]
        worksheet = raw_worksheet.iloc[header_row_index + 1 :].copy()
        worksheet.columns = header_values
        worksheet = worksheet.reset_index(drop=True)

        missing_columns = [column for column in WORKSHEET_COLUMNS if column not in worksheet.columns]
        if missing_columns:
            msg = (
                f"Worksheet '{self.worksheet_name}' is missing required columns: "
                f"{', '.join(missing_columns)}."
            )
            raise ValueError(msg)

        return worksheet

    def _find_header_row_index(self, raw_worksheet: pd.DataFrame) -> int:
        """Return the worksheet row index containing the expected headers."""
        expected = {column.casefold() for column in WORKSHEET_COLUMNS}
        for row_index in range(len(raw_worksheet.index)):
            row_values = {
                str(value).strip().casefold()
                for value in raw_worksheet.iloc[row_index].tolist()
                if not pd.isna(value) and str(value).strip()
            }
            if expected.issubset(row_values):
                return row_index

        msg = (
            f"Worksheet '{self.worksheet_name}' is missing required columns: "
            f"{', '.join(WORKSHEET_COLUMNS)}."
        )
        raise ValueError(msg)

    def _partition_records(
        self,
        website_records: list[DistributionRecord],
        worksheet_records: list[DistributionRecord],
    ) -> tuple[list[DistributionRecord], list[DistributionRecord], list[DistributionRecord]]:
        """Split website rows into missing, different, and matching groups."""
        worksheet_by_distros: dict[str, list[DistributionRecord]] = {}
        for record in worksheet_records:
            worksheet_by_distros.setdefault(record.distros, []).append(record)

        missing: list[DistributionRecord] = []
        different: list[DistributionRecord] = []
        matching: list[DistributionRecord] = []

        for record in website_records:
            candidates = worksheet_by_distros.get(record.distros, [])
            if not candidates:
                missing.append(record)
                continue

            if any(candidate.declared_ex_key() == record.declared_ex_key() for candidate in candidates):
                matching.append(record)
                continue

            different.append(record)

        return missing, different, matching

    def _write_sheet(
        self,
        missing_records: list[DistributionRecord],
        different_records: list[DistributionRecord],
    ) -> tuple[int, int]:
        """Write missing and changed distributions back to the workbook."""
        if not missing_records and not different_records:
            self.logger.info("No MSTY distribution rows need writing.")
            return 0, 0

        workbook = self.workbook_loader(self.excel_path)
        try:
            worksheet = workbook[self.worksheet_name]
        except KeyError as exc:
            workbook.close()
            self.logger.exception("Configured worksheet is missing during write.")
            raise ValueError(
                f"Worksheet '{self.worksheet_name}' was not found in {self.excel_path}.",
            ) from exc

        header_row_index, column_indexes = self._locate_distribution_table(worksheet)
        rows_updated = self._update_different_rows(
            worksheet,
            header_row_index,
            column_indexes,
            different_records,
        )
        rows_inserted = self._insert_missing_rows(
            worksheet,
            header_row_index,
            column_indexes,
            missing_records,
        )
        workbook.save(self.excel_path)
        workbook.close()
        self.logger.info(
            "Workbook update saved to %s with %s inserted rows and %s updated rows.",
            self.excel_path,
            rows_inserted,
            rows_updated,
        )
        return rows_inserted, rows_updated

    def _write_dc_pavula_sheet(
        self,
        website_records: list[DistributionRecord],
    ) -> tuple[
        int,
        int,
        list[dict[str, str | float]],
        list[dict[str, str | float | int]],
        list[dict[str, str | int]],
    ]:
        """Fill missing ROC% rows in the DC Pavula worksheet."""
        workbook = self.workbook_loader(self.excel_path)
        try:
            distribution_sheet = workbook[self.worksheet_name]
            dc_sheet = workbook[DC_PAVULA_WORKSHEET_NAME]
        except KeyError as exc:
            workbook.close()
            raise ValueError(f"Required worksheet missing in {self.excel_path}.") from exc

        _, distribution_columns = self._locate_distribution_table(distribution_sheet)
        distribution_row_map = self._build_distribution_row_map(
            distribution_sheet,
            distribution_columns,
        )
        dc_header_row = self._locate_dc_pavula_header_row(dc_sheet)
        dc_row_by_distribution = self._map_dc_rows_by_distribution_ref(dc_sheet, dc_header_row)

        pending_records = [
            record
            for record in website_records
            if record.roc_percent is not None and record.distribution_key() in distribution_row_map
        ]
        entries_to_log = [record.dc_entry() for record in pending_records]

        if not pending_records:
            workbook.close()
            return 0, 0, entries_to_log, [], []

        # Backup before we make any DC Pavula modifications (inserts + cell value writes).
        self._dc_backup_path = self._backup_excel_for_dc_pavula()
        self.logger.info(
            "DC Pavula pre-write backup: %s",
            self._dc_backup_path,
        )

        rows_inserted = 0
        rows_updated = 0
        skipped: list[dict[str, str | float | int]] = []
        merge_warnings: list[dict[str, str | int]] = []
        merge_warning_keys: set[tuple[int, str]] = set()
        for record in pending_records:
            distribution_row_index = distribution_row_map[record.distribution_key()]
            dc_row_index = dc_row_by_distribution.get(distribution_row_index)
            if dc_row_index is None:
                if not self.dc_pavula_insert_missing_rows:
                    skipped.append(
                        {
                            **record.dc_entry(),
                            "Distributions row": distribution_row_index,
                        },
                    )
                    self.logger.warning(
                        "DC Pavula: skipped ROC%% for Distributions row %s "
                        "(no DC row with Pay Date/Dist formula pointing at it). "
                        "Add the row in Excel or enable dc_pavula_insert_missing_rows.",
                        distribution_row_index,
                    )
                    continue
                dc_row_index = self._insert_dc_pavula_row(
                    dc_sheet,
                    dc_header_row,
                    dc_row_by_distribution,
                    distribution_row_index,
                )
                dc_row_by_distribution = self._shift_dc_row_map_after_insert(
                    dc_row_by_distribution,
                    dc_row_index,
                )
                dc_row_by_distribution[distribution_row_index] = dc_row_index
                rows_inserted += 1

            for coord in self._dc_pavula_merged_ranges_touching_row(dc_sheet, dc_row_index):
                key = (dc_row_index, coord)
                if key in merge_warning_keys:
                    continue
                merge_warning_keys.add(key)
                entry = {
                    "row": dc_row_index,
                    "merged_range": coord,
                    "dc_columns": "H:T",
                }
                merge_warnings.append(entry)
                self.logger.warning(
                    "DC Pavula: merged cell %s overlaps columns H–T on row %s; "
                    "openpyxl writes may not match Excel. Prefer unmerged data cells.",
                    coord,
                    dc_row_index,
                )

            roc_cell = dc_sheet.cell(row=dc_row_index, column=DC_PAVULA_ROC_COLUMN)
            preserved_format = roc_cell.number_format
            current_roc = self._normalize_optional_percentage(roc_cell.value)
            if current_roc != record.roc_percent:
                roc_cell.value = record.roc_percent
                if preserved_format:
                    roc_cell.number_format = preserved_format
                rows_updated += 1

        workbook.save(self.excel_path)
        workbook.close()
        return rows_inserted, rows_updated, entries_to_log, skipped, merge_warnings

    def _dc_pavula_merged_ranges_touching_row(
        self,
        worksheet: Worksheet,
        row_index: int,
    ) -> list[str]:
        """Return merged-range coordinates that intersect this row and columns H–T."""
        overlaps: list[str] = []
        start_col = DC_PAVULA_START_COLUMN
        end_col = DC_PAVULA_END_COLUMN
        for cell_range in worksheet.merged_cells.ranges:
            if cell_range.max_row < row_index or cell_range.min_row > row_index:
                continue
            if cell_range.max_col < start_col or cell_range.min_col > end_col:
                continue
            overlaps.append(str(cell_range.coord))
        return overlaps

    def _locate_distribution_table(
        self,
        worksheet: Worksheet,
    ) -> tuple[int, dict[str, int]]:
        """Find the header row and column indexes in the live worksheet."""
        expected = {column.casefold() for column in WORKSHEET_COLUMNS}
        for row_index in range(1, worksheet.max_row + 1):
            values_by_column: dict[str, int] = {}
            seen_values: set[str] = set()
            for column_index in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                lowered = text.casefold()
                seen_values.add(lowered)
                values_by_column[text] = column_index
            if expected.issubset(seen_values):
                return (
                    row_index,
                    {
                        column_name: values_by_column[column_name]
                        for column_name in WORKSHEET_COLUMNS
                    },
                )

        msg = (
            f"Worksheet '{self.worksheet_name}' is missing required columns: "
            f"{', '.join(WORKSHEET_COLUMNS)}."
        )
        raise ValueError(msg)

    def _update_different_rows(
        self,
        worksheet: Worksheet,
        header_row_index: int,
        column_indexes: dict[str, int],
        different_records: list[DistributionRecord],
    ) -> int:
        """Update existing rows whose Distros already exist but dates changed."""
        rows_updated = 0
        for record in different_records:
            row_index = self._find_row_by_distros(
                worksheet,
                header_row_index,
                column_indexes["Distros"],
                record.distros,
            )
            if row_index is None:
                continue
            self._write_record_to_row(worksheet, row_index, column_indexes, record)
            rows_updated += 1
        return rows_updated

    def _insert_missing_rows(
        self,
        worksheet: Worksheet,
        header_row_index: int,
        column_indexes: dict[str, int],
        missing_records: list[DistributionRecord],
    ) -> int:
        """Insert missing rows into the worksheet in declared-date order."""
        rows_inserted = 0
        for record in sorted(missing_records, key=self._record_sort_key):
            insertion_row = self._find_insertion_row(
                worksheet,
                header_row_index,
                column_indexes,
                record,
            )
            worksheet.insert_rows(insertion_row)
            template_row_index = insertion_row + 1 if insertion_row + 1 <= worksheet.max_row else insertion_row - 1
            if template_row_index >= 1:
                self._copy_row_format(worksheet, template_row_index, insertion_row)
            self._write_record_to_row(worksheet, insertion_row, column_indexes, record)
            rows_inserted += 1
        return rows_inserted

    def _find_row_by_distros(
        self,
        worksheet: Worksheet,
        header_row_index: int,
        distros_column_index: int,
        distros_value: str,
    ) -> int | None:
        """Find the first worksheet row matching the given distribution amount."""
        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=distros_column_index).value
            if value is None or str(value).strip() == "":
                continue
            try:
                normalized_value = self._normalize_distribution_value(value)
            except (TypeError, ValueError):
                continue
            if normalized_value == distros_value:
                return row_index
        return None

    def _find_insertion_row(
        self,
        worksheet: Worksheet,
        header_row_index: int,
        column_indexes: dict[str, int],
        record: DistributionRecord,
    ) -> int:
        """Return the row index where a missing record should be inserted."""
        declared_column_index = column_indexes["declared date"]
        target_date = self._parse_normalized_date(record.declared_date)
        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=declared_column_index).value
            if cell_value is None or str(cell_value).strip() == "":
                return row_index
            try:
                current_date = self._parse_excel_date(cell_value)
            except (TypeError, ValueError):
                continue
            if current_date > target_date:
                return row_index
        return worksheet.max_row + 1

    def _write_record_to_row(
        self,
        worksheet: Worksheet,
        row_index: int,
        column_indexes: dict[str, int],
        record: DistributionRecord,
    ) -> None:
        """Write one normalized distribution record into the worksheet row."""
        worksheet.cell(row=row_index, column=column_indexes["Distros"]).value = float(record.distros)
        worksheet.cell(
            row=row_index,
            column=column_indexes["declared date"],
        ).value = self._parse_normalized_date(record.declared_date).date()
        worksheet.cell(
            row=row_index,
            column=column_indexes["ex date"],
        ).value = self._parse_normalized_date(record.ex_date).date()
        worksheet.cell(
            row=row_index,
            column=column_indexes["record date"],
        ).value = self._parse_normalized_date(record.record_date).date()
        worksheet.cell(
            row=row_index,
            column=column_indexes["payable date"],
        ).value = self._parse_normalized_date(record.payable_date).date()

    def _copy_row_format(
        self,
        worksheet: Worksheet,
        source_row_index: int,
        target_row_index: int,
    ) -> None:
        """Copy cell styles from one worksheet row to another."""
        for column_index in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=source_row_index, column=column_index)
            target_cell = worksheet.cell(row=target_row_index, column=column_index)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format

        source_dimensions = worksheet.row_dimensions[source_row_index]
        target_dimensions = worksheet.row_dimensions[target_row_index]
        target_dimensions.height = source_dimensions.height
        target_dimensions.hidden = source_dimensions.hidden

    def _copy_dc_pavula_row(
        self,
        worksheet: Worksheet,
        source_row_index: int,
        target_row_index: int,
    ) -> None:
        """Copy formulas, values, and styles across the DC Pavula row block."""
        for column_index in range(DC_PAVULA_START_COLUMN, DC_PAVULA_END_COLUMN + 1):
            source_cell = worksheet.cell(row=source_row_index, column=column_index)
            target_cell = worksheet.cell(row=target_row_index, column=column_index)
            source_value = source_cell.value
            if isinstance(source_value, str) and source_value.startswith("="):
                target_cell.value = Translator(
                    source_value,
                    origin=source_cell.coordinate,
                ).translate_formula(target_cell.coordinate)
            else:
                target_cell.value = copy(source_value)

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format

        source_dimensions = worksheet.row_dimensions[source_row_index]
        target_dimensions = worksheet.row_dimensions[target_row_index]
        target_dimensions.height = source_dimensions.height
        target_dimensions.hidden = source_dimensions.hidden

    def _locate_dc_pavula_header_row(self, worksheet: Worksheet) -> int:
        """Find the header row in the DC Pavula worksheet."""
        expected = [value.casefold() for value in DC_PAVULA_HEADERS]
        for row_index in range(1, worksheet.max_row + 1):
            values = [
                str(worksheet.cell(row=row_index, column=column_index).value).strip().casefold()
                if worksheet.cell(row=row_index, column=column_index).value is not None
                else ""
                for column_index in range(
                    DC_PAVULA_START_COLUMN,
                    DC_PAVULA_START_COLUMN + len(DC_PAVULA_HEADERS),
                )
            ]
            if values == expected:
                return row_index
        msg = f"Worksheet '{DC_PAVULA_WORKSHEET_NAME}' is missing the expected header row."
        raise ValueError(msg)

    def _build_distribution_row_map(
        self,
        worksheet: Worksheet,
        column_indexes: dict[str, int],
    ) -> dict[tuple[str, str, str, str, str], int]:
        """Map distribution worksheet rows back to normalized record keys."""
        row_map: dict[tuple[str, str, str, str, str], int] = {}
        distros_column = column_indexes["Distros"]
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=distros_column).value
            if value is None or str(value).strip() == "":
                continue
            try:
                record = DistributionRecord(
                    distros=self._normalize_distribution_value(value),
                    declared_date=self._normalize_excel_date(
                        worksheet.cell(row=row_index, column=column_indexes["declared date"]).value,
                    ),
                    ex_date=self._normalize_excel_date(
                        worksheet.cell(row=row_index, column=column_indexes["ex date"]).value,
                    ),
                    record_date=self._normalize_excel_date(
                        worksheet.cell(row=row_index, column=column_indexes["record date"]).value,
                    ),
                    payable_date=self._normalize_excel_date(
                        worksheet.cell(row=row_index, column=column_indexes["payable date"]).value,
                    ),
                )
            except (TypeError, ValueError):
                continue
            row_map[record.distribution_key()] = row_index
        return row_map

    def _map_dc_rows_by_distribution_ref(
        self,
        worksheet: Worksheet,
        header_row_index: int,
    ) -> dict[int, int]:
        """Map referenced Distributions row numbers to DC Pavula row indexes."""
        mapping: dict[int, int] = {}
        for row_index in range(header_row_index + 1, worksheet.max_row + 1):
            pay_date_value = worksheet.cell(row=row_index, column=DC_PAVULA_START_COLUMN).value
            dist_value = worksheet.cell(row=row_index, column=DC_PAVULA_START_COLUMN + 1).value
            referenced_row = self._extract_distribution_row_reference(pay_date_value)
            if referenced_row is None:
                referenced_row = self._extract_distribution_row_reference(dist_value)
            if referenced_row is not None:
                mapping[referenced_row] = row_index
        return mapping

    def _extract_distribution_row_reference(self, value: Any) -> int | None:
        """Extract the referenced Distributions row number from a formula cell."""
        if not isinstance(value, str):
            return None
        match = DIST_SHEET_REF_PATTERN.search(value)
        if match is None:
            return None
        return int(match.group(1))

    def _insert_dc_pavula_row(
        self,
        worksheet: Worksheet,
        header_row_index: int,
        row_map: dict[int, int],
        distribution_row_index: int,
    ) -> int:
        """Insert a new DC Pavula row in the correct linked-distribution order."""
        later_rows = [
            (linked_distribution_row, dc_row_index)
            for linked_distribution_row, dc_row_index in row_map.items()
            if linked_distribution_row > distribution_row_index
        ]
        if later_rows:
            _, insertion_row = min(later_rows, key=lambda item: item[0])
            worksheet.insert_rows(insertion_row)
            self._copy_dc_pavula_row(worksheet, insertion_row + 1, insertion_row)
            return insertion_row

        if row_map:
            _, source_row_index = max(row_map.items(), key=lambda item: item[0])
            insertion_row = source_row_index + 1
            worksheet.insert_rows(insertion_row)
            self._copy_dc_pavula_row(worksheet, insertion_row - 1, insertion_row)
            return insertion_row

        insertion_row = header_row_index + 1
        worksheet.insert_rows(insertion_row)
        self._copy_dc_pavula_row(worksheet, insertion_row + 1, insertion_row)
        return insertion_row

    def _shift_dc_row_map_after_insert(
        self,
        row_map: dict[int, int],
        insertion_row: int,
    ) -> dict[int, int]:
        """Shift DC row indexes after inserting a new worksheet row."""
        return {
            linked_distribution_row: (
                dc_row_index + 1 if dc_row_index >= insertion_row else dc_row_index
            )
            for linked_distribution_row, dc_row_index in row_map.items()
        }

    def _record_sort_key(self, record: DistributionRecord) -> tuple[datetime, str]:
        """Sort records by declared date, then distribution amount."""
        return (self._parse_normalized_date(record.declared_date), record.distros)

    def _row_has_blank_dates(self, row: pd.Series) -> bool:
        """Skip worksheet rows that do not contain a full distribution date set."""
        return any(pd.isna(row[column]) or str(row[column]).strip() == "" for column in WORKSHEET_COLUMNS[1:])

    def _normalize_date(self, value: Any) -> str:
        """Convert website dates from mm/dd/yyyy into dd/Mon/yyyy."""
        parsed = datetime.strptime(str(value).strip(), "%m/%d/%Y")
        return parsed.strftime(DATE_OUTPUT_FORMAT)

    def _normalize_excel_date(self, value: Any) -> str:
        """Normalize worksheet dates into the expected comparison format."""
        timestamp = self._parse_excel_date(value)
        return timestamp.strftime(DATE_OUTPUT_FORMAT)

    def _normalize_distribution_value(self, value: Any) -> str:
        """Format distribution values without currency symbols, preserving decimals."""
        numeric = pd.to_numeric(str(value).replace("$", "").strip(), errors="raise")
        return f"{float(numeric):.4f}"

    def _normalize_roc_percent(self, value: Any) -> float | None:
        """Normalize ROC percentage text into a numeric percent value."""
        text = str(value).replace("%", "").strip()
        if text == "" or text.lower() == "nan":
            return None
        numeric = pd.to_numeric(text, errors="raise")
        return round(float(numeric), 2)

    def _normalize_optional_percentage(self, value: Any) -> float | None:
        """Normalize an existing worksheet ROC value for comparison."""
        if value is None or str(value).strip() == "":
            return None
        numeric = pd.to_numeric(value, errors="raise")
        return round(float(numeric), 2)

    def _parse_normalized_date(self, value: str) -> datetime:
        """Parse a normalized worksheet date string."""
        return datetime.strptime(value, DATE_PARSE_FORMAT)

    def _parse_excel_date(self, value: Any) -> datetime:
        """Parse a worksheet cell date value into a datetime."""
        if isinstance(value, str) and value.strip().startswith("="):
            msg = "Formula cells cannot be parsed as fixed worksheet dates."
            raise ValueError(msg)
        timestamp = pd.to_datetime(value, errors="raise")
        return timestamp.to_pydatetime()
