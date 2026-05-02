"""Unit tests for Sharesight CAPITAL_RETURN trades skill."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from openclaw.skills.sharesight_trades.skill import SharesightTradesSkill, TradeRecord


def _trade(
    *,
    trade_id: int,
    transaction_date,
    roc: float,
    fx: float,
    holding_id: int = 1234,
    price_value: float | None = None,
) -> TradeRecord:
    raw: dict[str, object] = {
        "id": trade_id,
        "transaction_date": transaction_date.isoformat() if transaction_date else None,
        "transaction_type": "CAPITAL_RETURN",
        "holding_id": holding_id,
        "capital_return_value": roc,
        "exchange_rate": fx,
    }
    if price_value is not None:
        raw["price"] = price_value
    return TradeRecord(
        id=trade_id,
        company_event_id=None,
        transaction_date=transaction_date,
        transaction_type="CAPITAL_RETURN",
        holding_id=holding_id,
        unique_identifier=None,
        roc_value=roc,
        exchange_rate_value=fx,
        price_value=price_value,
        raw=raw,
    )


def create_workbook_fixture(workbook_path: Path) -> None:
    """Create workbook with the CS FY2526 columns needed by this skill."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CS FY2526"
    worksheet.cell(row=6, column=8).value = "CS distributions"
    worksheet.cell(row=7, column=8).value = "Pay Date"
    worksheet.cell(row=7, column=10).value = "ROC%"
    worksheet.cell(row=7, column=12).value = "ROC $"
    worksheet.cell(row=7, column=14).value = "Gross Amt"
    worksheet.cell(row=7, column=20).value = "Exchange Rate"

    worksheet.cell(row=8, column=8).value = datetime(2026, 1, 23)
    worksheet.cell(row=8, column=10).value = 94.69
    worksheet.cell(row=8, column=12).value = 12.34
    worksheet.cell(row=8, column=14).value = 571.47
    worksheet.cell(row=8, column=20).value = 1.45

    workbook.save(workbook_path)
    workbook.close()


def add_workbook_row(
    workbook_path: Path,
    *,
    row_index: int,
    pay_date: datetime,
    roc_percent: float,
    roc_amount,
    gross_amount: float,
    exchange_rate,
) -> None:
    """Append or overwrite one worksheet row for test scenarios."""
    from openpyxl import load_workbook as _load_workbook

    wb = _load_workbook(workbook_path)
    ws = wb["CS FY2526"]
    ws.cell(row=row_index, column=8).value = pay_date
    ws.cell(row=row_index, column=10).value = roc_percent
    ws.cell(row=row_index, column=12).value = roc_amount
    ws.cell(row=row_index, column=14).value = gross_amount
    ws.cell(row=row_index, column=20).value = exchange_rate
    wb.save(workbook_path)
    wb.close()


class FakeApiClient:
    """Minimal API test double for reconcile."""

    def __init__(self) -> None:
        self.created_payloads: list[dict[str, object]] = []
        self.updated: list[tuple[int, dict[str, object]]] = []
        self.deleted_ids: list[int] = []
        self.existing_trades: list[TradeRecord] = []
        self.list_trades_calls: list[tuple[int, object, object]] = []
        self.closed = False
        self._next_id = 9001

    def resolve_portfolio_id(self, portfolio_name: str) -> int:
        """Return fixed fake portfolio id."""
        return 1201759

    def create_trade(self, payload: dict[str, object]) -> TradeRecord:
        """Record create payload and return fake created trade."""
        self.created_payloads.append(payload)
        tid = self._next_id
        self._next_id += 1
        trade = payload["trade"]
        td = datetime.strptime(str(trade["transaction_date"]), "%Y-%m-%d").date()
        roc = float(trade["capital_return_value"])
        fx = float(trade["exchange_rate"])
        return _trade(trade_id=tid, transaction_date=td, roc=roc, fx=fx, price_value=0.0)

    def update_trade(self, trade_id: int, payload: dict[str, object]) -> TradeRecord:
        """Record update."""
        self.updated.append((trade_id, payload))
        trade = payload["trade"]
        td = datetime.strptime(str(trade["transaction_date"]), "%Y-%m-%d").date()
        roc = float(trade["capital_return_value"])
        fx = float(trade["exchange_rate"])
        pv = trade.get("price")
        price_out = None if pv in (None, "") else float(pv)
        return _trade(trade_id=trade_id, transaction_date=td, roc=roc, fx=fx, price_value=price_out)

    def delete_trade(self, trade_id: int) -> None:
        """Record delete."""
        self.deleted_ids.append(trade_id)

    def list_trades(self, portfolio_id: int, *, start_date=None, end_date=None) -> list[TradeRecord]:
        """Return configured existing trades."""
        self.list_trades_calls.append((portfolio_id, start_date, end_date))
        return list(self.existing_trades)

    def close(self) -> None:
        """Mark fake client closed."""
        self.closed = True


class TestSharesightTradesSkill(unittest.TestCase):
    """Coverage for worksheet mapping and reconcile behavior."""

    def test_dry_run_can_be_overridden_at_runtime(self) -> None:
        """Runtime dry_run=True avoids API write calls."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=False,
                api_factory=lambda: api,
            )

            result = skill.run(dry_run=True)

        self.assertTrue(result["dry_run"])
        self.assertEqual(result["rows_read"], 1)
        self.assertEqual(result["created_count"], 0)
        self.assertEqual(result["deleted_count"], 0)
        self.assertEqual(result["updated_count"], 0)
        self.assertEqual(result["dry_run_summary"]["to_add_count"], 1)
        self.assertEqual(len(result["reconcile_add"]), 1)
        self.assertEqual(len(result["reconcile_noop"]), 0)
        payload = result["reconcile_add"][0]["create_payload"]["trade"]
        self.assertEqual(payload["transaction_type"], "CAPITAL_RETURN")
        self.assertEqual(payload["transaction_date"], "2026-01-23")
        self.assertEqual(payload["paid_on"], "2026-01-23")
        self.assertEqual(payload["price"], 0.0)
        self.assertEqual(payload["capital_return_value"], 12.34)
        self.assertEqual(payload["exchange_rate"], 1.45)
        self.assertEqual(payload["comments"], "94.69% 23-Jan ROC. Gross Amt $571.47")
        self.assertEqual(payload["state"], "confirmed")
        self.assertEqual(api.created_payloads, [])
        self.assertEqual(api.deleted_ids, [])
        self.assertEqual(api.updated, [])
        self.assertTrue(api.closed)

    def test_run_creates_when_no_existing_trade(self) -> None:
        """Non dry-run creates trade when sheet has row and API has none."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=False,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertFalse(result["dry_run"])
        self.assertEqual(result["created_count"], 1)
        self.assertEqual(result["deleted_count"], 0)
        self.assertEqual(result["updated_count"], 0)
        self.assertEqual(result["noop_count"], 0)
        self.assertEqual(len(api.created_payloads), 1)
        self.assertEqual(api.created_payloads[0]["trade"]["state"], "confirmed")

    def test_dry_run_noop_when_trade_matches_sheet(self) -> None:
        """Existing trade with same ROC and FX is noop."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            d = datetime(2026, 1, 23).date()
            api.existing_trades = [_trade(trade_id=20995465, transaction_date=d, roc=12.34, fx=1.45)]
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["dry_run_summary"]["noop_count"], 1)
        self.assertEqual(result["dry_run_summary"]["to_add_count"], 0)
        self.assertEqual(result["dry_run_summary"]["to_delete_count"], 0)
        self.assertEqual(result["dry_run_summary"]["to_update_count"], 0)
        self.assertEqual(api.created_payloads, [])

    def test_dry_run_update_when_roc_differs(self) -> None:
        """Dry run lists update when ROC differs."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            d = datetime(2026, 1, 23).date()
            api.existing_trades = [_trade(trade_id=20995465, transaction_date=d, roc=99.0, fx=1.45)]
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["dry_run_summary"]["to_update_count"], 1)
        self.assertEqual(len(result["reconcile_update"]), 1)
        self.assertEqual(result["reconcile_update"][0]["trade_id"], 20995465)

    def test_dry_run_update_when_roc_fx_match_but_api_price_nonzero(self) -> None:
        """Matching ROC/FX with duplicate non-zero API price still triggers update to clear price."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            d = datetime(2026, 1, 23).date()
            api.existing_trades = [
                _trade(trade_id=20995465, transaction_date=d, roc=12.34, fx=1.45, price_value=12.34),
            ]
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=True,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["dry_run_summary"]["to_update_count"], 1)
        self.assertEqual(result["dry_run_summary"]["noop_count"], 0)
        self.assertEqual(len(result["reconcile_update"]), 1)
        item = result["reconcile_update"][0]
        self.assertEqual(item["trade_id"], 20995465)
        self.assertEqual(item["existing_price"], 12.34)

    def test_live_update_and_delete_orphan(self) -> None:
        """Update mismatched trade; delete orphan date not on sheet."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            d = datetime(2026, 1, 23).date()
            orphan = datetime(2025, 6, 1).date()
            api.existing_trades = [
                _trade(trade_id=100, transaction_date=d, roc=99.0, fx=1.45),
                _trade(trade_id=200, transaction_date=orphan, roc=5.0, fx=1.0),
            ]
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=False,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["updated_count"], 1)
        self.assertEqual(result["deleted_count"], 1)
        self.assertEqual(result["created_count"], 0)
        self.assertIn(200, api.deleted_ids)
        self.assertEqual(api.updated[0][0], 100)

    def test_duplicate_trades_same_date_deletes_extras(self) -> None:
        """Second trade on same pay date is deleted (keep lowest id)."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            api = FakeApiClient()
            d = datetime(2026, 1, 23).date()
            api.existing_trades = [
                _trade(trade_id=50, transaction_date=d, roc=12.34, fx=1.45),
                _trade(trade_id=51, transaction_date=d, roc=12.34, fx=1.45),
            ]
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=False,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["noop_count"], 1)
        self.assertEqual(result["deleted_count"], 1)
        self.assertIn(51, api.deleted_ids)
        self.assertNotIn(50, api.deleted_ids)

    def test_run_skips_invalid_rows_with_non_positive_roc_amount(self) -> None:
        """Rows with non-positive ROC amount are skipped; valid row still reconciles."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            add_workbook_row(
                workbook_path,
                row_index=9,
                pay_date=datetime(2026, 2, 1),
                roc_percent=90.0,
                roc_amount=0,
                gross_amount=100.0,
                exchange_rate=1.2,
            )
            api = FakeApiClient()
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=False,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["invalid_rows_skipped_count"], 1)
        self.assertEqual(result["created_count"], 1)

    def test_run_skips_invalid_rows_with_non_positive_exchange_rate(self) -> None:
        """Rows with non-positive exchange rate are skipped."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Personal CashFlow.xlsx"
            create_workbook_fixture(workbook_path)
            add_workbook_row(
                workbook_path,
                row_index=9,
                pay_date=datetime(2026, 2, 1),
                roc_percent=90.0,
                roc_amount=11.5,
                gross_amount=100.0,
                exchange_rate=0,
            )
            api = FakeApiClient()
            skill = SharesightTradesSkill(
                excel_path=workbook_path,
                worksheet_name="CS FY2526",
                portfolio_name="DC Pavula",
                holding_id=1234,
                client_id="client-id",
                client_secret="client-secret",
                dry_run=False,
                api_factory=lambda: api,
            )

            result = skill.run()

        self.assertEqual(result["invalid_rows_skipped_count"], 1)
        self.assertEqual(result["created_count"], 1)


if __name__ == "__main__":
    unittest.main()
