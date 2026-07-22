"""Sharesight trade upload skill for CAPITAL_RETURN rows."""

from __future__ import annotations

from base64 import b64encode
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
import json
import os
from pathlib import Path
from typing import Any, Callable, Protocol
from urllib import error, parse, request

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from openclaw.skills._base.skill_base import SkillBase
from openclaw.skills.sharesight_trades.log_writer import ObsidianRunLogWriter

DEFAULT_API_BASE_URL = "https://api.sharesight.com/api/v2"
DEFAULT_TOKEN_URL = "https://api.sharesight.com/oauth2/token"
DEFAULT_EXCEL_PATH = "~/OneDrive/Documents/Finance/Personal CashFlow.xlsx"
DEFAULT_EXCEL_PATH_ENV_VAR = "OPENCLAW_SHARESIGHT_SYNC_EXCEL_PATH"
DEFAULT_WORKSHEET_NAME = "CS FY2627"
DEFAULT_PORTFOLIO_NAME = "DC Pavula"
DEFAULT_CLIENT_ID_ENV = "SHARESIGHT_CLIENT_ID"
DEFAULT_CLIENT_SECRET_ENV = "SHARESIGHT_CLIENT_SECRET"
DEFAULT_TRANSACTION_TYPE = "CAPITAL_RETURN"
DEFAULT_CREATED_STATE = "confirmed"
DEFAULT_CONFIRMED_STATE = "confirmed"
DATE_OUTPUT_FORMAT = "%d/%m/%Y"
API_DATE_FORMAT = "%Y-%m-%d"
HEADER_START_COLUMN = 8
REQUIRED_HEADERS = {
    "Pay Date": 8,
    "ROC%": 10,
    "ROC $": 12,
    "Gross Amt": 14,
}
EXCHANGE_RATE_COLUMN = 20  # column T
FLOAT_COMPARE_EPSILON = 1e-6
DEFAULT_LOG_ENV = "local"
DEFAULT_LOG_OPERATOR = "unknown"
DEFAULT_OBSIDIAN_USER = "bobbyd"


@dataclass(slots=True)
class WorksheetEntry:
    """One worksheet row converted to trade fields."""

    row_index: int
    pay_date: date
    roc_percent: str
    roc_amount: str
    gross_amount: str
    exchange_rate: str

    def comment_text(self) -> str:
        """Build the Sharesight comment text for this row."""
        short_date = f"{self.pay_date.day}-{self.pay_date.strftime('%b')}"
        gross_value = format_currency(self.gross_amount)
        return f"{self.roc_percent}% {short_date} ROC. Gross Amt ${gross_value}"


@dataclass(slots=True)
class TradeRecord:
    """A Sharesight trade returned by the API."""

    id: int | None
    company_event_id: int | None
    transaction_date: date | None
    transaction_type: str
    holding_id: int | None
    unique_identifier: str | None
    roc_value: float | None
    exchange_rate_value: float | None
    price_value: float | None
    raw: dict[str, Any]


class SharesightTradeApi(Protocol):
    """API-facing contract for trade reconcile operations."""

    def resolve_portfolio_id(self, portfolio_name: str) -> int:
        """Resolve a portfolio ID from its visible name."""

    def create_trade(self, payload: dict[str, Any]) -> TradeRecord:
        """Create one trade."""

    def update_trade(self, trade_id: int, payload: dict[str, Any]) -> TradeRecord:
        """Update one trade (PUT)."""

    def delete_trade(self, trade_id: int) -> None:
        """Delete one trade (DELETE)."""

    def list_trades(
        self,
        portfolio_id: int,
        *,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[TradeRecord]:
        """List trades for one portfolio."""

    def close(self) -> None:
        """Close any API resources."""


class SharesightTradesSkill(SkillBase):
    """Reconcile Sharesight CAPITAL_RETURN trades with Excel (source of truth)."""

    def __init__(
        self,
        config_path: str | Path | None = None,
        *,
        api_base_url: str | None = None,
        token_url: str | None = None,
        excel_path: str | Path | None = None,
        excel_path_env_var: str | None = None,
        worksheet_name: str | None = None,
        portfolio_name: str | None = None,
        holding_id: int | None = None,
        client_id: str | None = None,
        client_secret: str | None = None,
        client_id_env: str | None = None,
        client_secret_env: str | None = None,
        dry_run: bool | None = None,
        transaction_type: str | None = None,
        created_state: str | None = None,
        confirmed_state: str | None = None,
        unique_identifier_prefix: str | None = None,
        workbook_loader: Callable[..., Any] | None = None,
        api_factory: Callable[[], SharesightTradeApi] | None = None,
    ) -> None:
        """Initialize the trade skill."""
        super().__init__(config_path)
        self.api_base_url = str(self.config.get("api_base_url", api_base_url or DEFAULT_API_BASE_URL)).rstrip("/")
        self.token_url = str(self.config.get("token_url", token_url or DEFAULT_TOKEN_URL))
        self.excel_path_env_var = str(
            excel_path_env_var or self.config.get("excel_path_env_var", DEFAULT_EXCEL_PATH_ENV_VAR),
        )
        self.excel_path = self._resolve_excel_path(excel_path)
        self.worksheet_name = str(self.config.get("worksheet_name", worksheet_name or DEFAULT_WORKSHEET_NAME))
        self.portfolio_name = str(self.config.get("portfolio_name", portfolio_name or DEFAULT_PORTFOLIO_NAME))
        resolved_holding_id = self.config.get("holding_id", holding_id)
        self.holding_id = None if resolved_holding_id in (None, "") else int(resolved_holding_id)
        self.client_id = client_id or self._read_secret(
            env_name=str(self.config.get("client_id_env", client_id_env or DEFAULT_CLIENT_ID_ENV)),
            label="Sharesight client ID",
        )
        self.client_secret = client_secret or self._read_secret(
            env_name=str(self.config.get("client_secret_env", client_secret_env or DEFAULT_CLIENT_SECRET_ENV)),
            label="Sharesight client secret",
        )
        configured_dry_run = self._coerce_optional_bool(
            dry_run if dry_run is not None else self.config.get("dry_run"),
        )
        self.dry_run = False if configured_dry_run is None else configured_dry_run
        self.transaction_type = str(
            self.config.get("transaction_type", transaction_type or DEFAULT_TRANSACTION_TYPE),
        ).upper()
        self.created_state = str(self.config.get("created_state", created_state or DEFAULT_CREATED_STATE))
        self.confirmed_state = str(self.config.get("confirmed_state", confirmed_state or DEFAULT_CONFIRMED_STATE))
        self.unique_identifier_prefix = str(
            self.config.get("unique_identifier_prefix", unique_identifier_prefix or "OPENCLAW-CS-ROC"),
        )
        self.workbook_loader = workbook_loader or load_workbook
        self._api_factory = api_factory or (
            lambda: SharesightTradeApiClient(
                api_base_url=self.api_base_url,
                token_url=self.token_url,
                client_id=self.client_id,
                client_secret=self.client_secret,
            )
        )
        if self.holding_id is None:
            raise ValueError("holding_id is required for sharesight_trades skill configuration.")
        if self.transaction_type != DEFAULT_TRANSACTION_TYPE:
            raise ValueError("sharesight_trades only supports transaction_type CAPITAL_RETURN.")

    def _write_obsidian_run_log(self, result_payload: dict[str, Any]) -> None:
        """Write an Obsidian run log when explicitly enabled in config."""
        if not bool(self.config.get("obsidian_log_enabled", False)):
            return

        configured_logs_dir = self.config.get("obsidian_log_dir")
        if not configured_logs_dir:
            self.logger.warning(
                "Obsidian logging enabled, but `obsidian_log_dir` is not configured.",
            )
            return

        log_writer = ObsidianRunLogWriter(
            logs_dir=Path(str(configured_logs_dir)).expanduser(),
            operator=str(self.config.get("obsidian_log_operator", DEFAULT_LOG_OPERATOR)),
            environment=str(self.config.get("obsidian_log_environment", DEFAULT_LOG_ENV)),
            obsidian_user=str(self.config.get("obsidian_log_user", DEFAULT_OBSIDIAN_USER)),
            config_path=str(
                self.config.get("obsidian_log_config_path", "skills/sharesight_trades/config.yaml"),
            ),
            api_base_url=self.api_base_url,
            env_var_name=self.excel_path_env_var,
        )

        output_path = log_writer.write_log(
            result_payload,
            excel_path_resolved=str(self.excel_path),
            env_override_used=bool(os.getenv(self.excel_path_env_var)),
            command_used=str(
                self.config.get(
                    "obsidian_log_command_used",
                    "python - <<'PY' ... SharesightTradesSkill(...).run() ... PY",
                ),
            ),
            notes=["Generated automatically by SharesightTradesSkill."],
        )
        self.logger.info("Wrote Obsidian run log to %s.", output_path)

    def run(self, dry_run: bool | str | int | None = None) -> dict[str, Any]:
        """Reconcile trades: add missing, update mismatched ROC/FX, delete orphans."""
        resolved_dry_run = self._resolve_runtime_dry_run(dry_run)
        self.logger.info(
            "Starting Sharesight trade reconcile for portfolio %s using %s [%s].",
            self.portfolio_name,
            self.excel_path,
            self.worksheet_name,
        )
        entries, skipped_rows = self._read_worksheet_entries()
        api = self._api_factory()

        invalid_rows: list[dict[str, Any]] = []
        all_sheet_dates: set[date] = {e.pay_date for e in entries}
        desired_by_date: dict[date, WorksheetEntry] = {}

        for entry in entries:
            invalid_reason = self._validate_entry(entry)
            if invalid_reason is not None:
                invalid_rows.append(
                    {
                        "row": entry.row_index,
                        "pay_date": entry.pay_date.isoformat(),
                        "reason": invalid_reason,
                    },
                )
                print(
                    "SKIP INVALID "
                    f"row={entry.row_index} date={entry.pay_date.isoformat()} reason={invalid_reason}",
                )
                continue
            desired_by_date[entry.pay_date] = entry

        reconcile_add: list[dict[str, Any]] = []
        reconcile_update: list[dict[str, Any]] = []
        reconcile_delete: list[dict[str, Any]] = []
        reconcile_noop: list[dict[str, Any]] = []

        created_count = 0
        updated_count = 0
        deleted_count = 0
        noop_count = 0
        created_trade_ids: list[int] = []
        updated_trade_ids: list[int] = []
        deleted_trade_ids: list[int] = []
        existing_trades_count = 0

        try:
            if not all_sheet_dates:
                noop_count = 0
                early_result = self._build_result(
                    resolved_dry_run=resolved_dry_run,
                    entries=entries,
                    skipped_rows=skipped_rows,
                    invalid_rows=invalid_rows,
                    existing_trades_count=0,
                    reconcile_add=reconcile_add,
                    reconcile_update=reconcile_update,
                    reconcile_delete=reconcile_delete,
                    reconcile_noop=reconcile_noop,
                    created_count=0,
                    updated_count=0,
                    deleted_count=0,
                    noop_count=noop_count,
                    created_trade_ids=[],
                    updated_trade_ids=[],
                    deleted_trade_ids=[],
                )
                self._write_obsidian_run_log(early_result)
                return early_result

            portfolio_id = api.resolve_portfolio_id(self.portfolio_name)
            start_date = min(all_sheet_dates)
            end_date = max(all_sheet_dates)
            existing_trades = api.list_trades(
                portfolio_id,
                start_date=start_date,
                end_date=end_date,
            )
            existing_trades_count = len(existing_trades)

            managed = [
                t
                for t in existing_trades
                if t.id is not None
                and t.transaction_date is not None
                and t.holding_id == int(self.holding_id)
                and t.transaction_type == self.transaction_type
            ]

            by_date: dict[date, list[TradeRecord]] = {}
            for t in managed:
                by_date.setdefault(t.transaction_date, []).append(t)
            for group in by_date.values():
                group.sort(key=lambda x: int(x.id or 0))

            prefix = "DRY-RUN " if resolved_dry_run else ""

            for t in managed:
                if t.transaction_date not in all_sheet_dates:
                    reconcile_delete.append(
                        {
                            "trade_id": t.id,
                            "transaction_date": t.transaction_date.isoformat(),
                            "reason": "pay_date not present in spreadsheet",
                            "existing_price": t.price_value,
                        },
                    )
                    print(
                        f"{prefix}DELETE trade_id={t.id} date={t.transaction_date.isoformat()} "
                        "reason=not_in_spreadsheet"
                        + self._existing_price_print_suffix(t.price_value),
                    )

            for pay_d, entry in desired_by_date.items():
                trades_at = list(by_date.get(pay_d, []))
                extras = trades_at[1:]
                for ex in extras:
                    reconcile_delete.append(
                        {
                            "trade_id": ex.id,
                            "transaction_date": pay_d.isoformat(),
                            "reason": "duplicate trade for same pay_date",
                            "existing_price": ex.price_value,
                        },
                    )
                    print(
                        f"{prefix}DELETE trade_id={ex.id} date={pay_d.isoformat()} reason=duplicate"
                        + self._existing_price_print_suffix(ex.price_value),
                    )

                primary = trades_at[0] if trades_at else None
                if primary is None:
                    create_payload = self._build_create_payload(portfolio_id, entry)
                    reconcile_add.append(
                        {
                            "row": entry.row_index,
                            "pay_date": pay_d.isoformat(),
                            "create_payload": create_payload,
                        },
                    )
                    print(
                        f"{prefix}ADD row={entry.row_index} date={pay_d.isoformat()} "
                        f"roc={entry.roc_amount} fx={entry.exchange_rate}",
                    )
                elif self._trade_matches_entry(primary, entry):
                    reconcile_noop.append(
                        {
                            "row": entry.row_index,
                            "pay_date": pay_d.isoformat(),
                            "trade_id": primary.id,
                            "existing_roc": primary.roc_value,
                            "existing_exchange_rate": primary.exchange_rate_value,
                            "existing_price": primary.price_value,
                        },
                    )
                    print(
                        f"{prefix}NOOP row={entry.row_index} trade_id={primary.id} date={pay_d.isoformat()}"
                        + self._existing_price_print_suffix(primary.price_value),
                    )
                else:
                    update_payload = self._build_update_payload(portfolio_id, entry)
                    reconcile_update.append(
                        {
                            "row": entry.row_index,
                            "pay_date": pay_d.isoformat(),
                            "trade_id": primary.id,
                            "existing_roc": primary.roc_value,
                            "existing_exchange_rate": primary.exchange_rate_value,
                            "existing_price": primary.price_value,
                            "desired_roc": float(entry.roc_amount),
                            "desired_exchange_rate": float(entry.exchange_rate),
                            "update_payload": update_payload,
                        },
                    )
                    print(
                        f"{prefix}UPDATE trade_id={primary.id} row={entry.row_index} date={pay_d.isoformat()} "
                        f"roc {primary.roc_value}->{entry.roc_amount} fx {primary.exchange_rate_value}->{entry.exchange_rate}"
                        + self._existing_price_print_suffix(primary.price_value),
                    )

            noop_count = len(reconcile_noop)

            if not resolved_dry_run:
                for item in reconcile_delete:
                    tid = int(item["trade_id"])
                    api.delete_trade(tid)
                    deleted_count += 1
                    deleted_trade_ids.append(tid)

                for item in reconcile_update:
                    tid = int(item["trade_id"])
                    api.update_trade(tid, item["update_payload"])
                    updated_count += 1
                    updated_trade_ids.append(tid)

                for item in reconcile_add:
                    created = api.create_trade(item["create_payload"])
                    created_count += 1
                    if created.id is not None:
                        created_trade_ids.append(created.id)

        finally:
            api.close()

        final_result = self._build_result(
            resolved_dry_run=resolved_dry_run,
            entries=entries,
            skipped_rows=skipped_rows,
            invalid_rows=invalid_rows,
            existing_trades_count=existing_trades_count,
            reconcile_add=reconcile_add,
            reconcile_update=reconcile_update,
            reconcile_delete=reconcile_delete,
            reconcile_noop=reconcile_noop,
            created_count=created_count if not resolved_dry_run else 0,
            updated_count=updated_count if not resolved_dry_run else 0,
            deleted_count=deleted_count if not resolved_dry_run else 0,
            noop_count=noop_count,
            created_trade_ids=created_trade_ids,
            updated_trade_ids=updated_trade_ids,
            deleted_trade_ids=deleted_trade_ids,
        )
        self._write_obsidian_run_log(final_result)
        return final_result

    def _build_result(
        self,
        *,
        resolved_dry_run: bool,
        entries: list[WorksheetEntry],
        skipped_rows: list[dict[str, str]],
        invalid_rows: list[dict[str, Any]],
        existing_trades_count: int,
        reconcile_add: list[dict[str, Any]],
        reconcile_update: list[dict[str, Any]],
        reconcile_delete: list[dict[str, Any]],
        reconcile_noop: list[dict[str, Any]],
        created_count: int,
        updated_count: int,
        deleted_count: int,
        noop_count: int,
        created_trade_ids: list[int],
        updated_trade_ids: list[int],
        deleted_trade_ids: list[int],
    ) -> dict[str, Any]:
        """Assemble run() response dict."""
        dry_add = list(reconcile_add)
        dry_update = list(reconcile_update)
        dry_delete = list(reconcile_delete)
        dry_noop = list(reconcile_noop)
        return {
            "status": "success",
            "dry_run": resolved_dry_run,
            "api_base_url": self.api_base_url,
            "workbook_path": str(self.excel_path),
            "worksheet_name": self.worksheet_name,
            "portfolio_name": self.portfolio_name,
            "holding_id": self.holding_id,
            "transaction_type": self.transaction_type,
            "rows_read": len(entries),
            "rows_skipped": len(skipped_rows),
            "skipped_rows": skipped_rows,
            "invalid_rows_skipped_count": len(invalid_rows),
            "invalid_rows": invalid_rows,
            "existing_trades_fetched": existing_trades_count,
            "reconcile_add": dry_add,
            "reconcile_update": dry_update,
            "reconcile_delete": dry_delete,
            "reconcile_noop": dry_noop,
            "created_count": created_count,
            "updated_count": updated_count,
            "deleted_count": deleted_count,
            "noop_count": noop_count,
            "matched_and_skipped_count": noop_count,
            "created_trade_ids": created_trade_ids,
            "updated_trade_ids": updated_trade_ids,
            "deleted_trade_ids": deleted_trade_ids,
            "confirmed_count": created_count,
            "confirmed_trade_ids": list(created_trade_ids),
            "dry_run_payloads": dry_add,
            "dry_run_new_trades": dry_add,
            "dry_run_matches": dry_noop,
            "dry_run_summary": {
                "worksheet_rows_total": len(entries),
                "existing_trades_fetched": existing_trades_count,
                "to_add_count": len(dry_add),
                "to_update_count": len(dry_update),
                "to_delete_count": len(dry_delete),
                "noop_count": len(dry_noop),
                "invalid_rows_skipped_count": len(invalid_rows),
                "matched_and_skipped_count": noop_count,
            },
        }

    def _trade_matches_entry(self, trade: TradeRecord, entry: WorksheetEntry) -> bool:
        """Return True if API trade ROC and FX match worksheet and API price field is ~zero or absent."""
        try:
            want_roc = float(Decimal(str(entry.roc_amount).strip()))
            want_fx = float(Decimal(str(entry.exchange_rate).strip()))
        except (InvalidOperation, ValueError):
            return False
        if not (
            self._floats_match(trade.roc_value, want_roc)
            and self._floats_match(trade.exchange_rate_value, want_fx)
        ):
            return False
        return self._api_price_is_normalized(trade)

    def _api_price_is_normalized(self, trade: TradeRecord) -> bool:
        """True when Sharesight `price` is missing or ~0 (ROC-only semantics)."""
        if trade.price_value is None:
            return True
        return abs(trade.price_value) <= FLOAT_COMPARE_EPSILON

    @staticmethod
    def _existing_price_print_suffix(price_value: float | None) -> str:
        """Append API price to stdout when present."""
        if price_value is None:
            return ""
        return f" existing_price={price_value}"

    def _floats_match(self, left: float | None, right: float | None) -> bool:
        """Approximate float equality."""
        if left is None or right is None:
            return left == right
        return abs(left - right) <= FLOAT_COMPARE_EPSILON

    def _resolve_runtime_dry_run(self, dry_run: bool | str | int | None) -> bool:
        """Resolve configured dry-run with optional runtime override."""
        override = self._coerce_optional_bool(dry_run)
        return self.dry_run if override is None else override

    def _validate_entry(self, entry: WorksheetEntry) -> str | None:
        """Return validation error message when row cannot be posted."""
        roc_amount = self._to_decimal_or_none(entry.roc_amount)
        if roc_amount is None or roc_amount <= 0:
            return "capital_return_value must be greater than zero"
        exchange_rate = self._to_decimal_or_none(entry.exchange_rate)
        if exchange_rate is None or exchange_rate <= 0:
            return "exchange_rate must be greater than zero"
        return None

    def _resolve_excel_path(self, excel_path: str | Path | None) -> Path:
        """Resolve workbook path from constructor, env var, then config."""
        if excel_path is not None:
            return Path(excel_path).expanduser()
        env_excel_path = os.getenv(self.excel_path_env_var)
        if env_excel_path:
            return Path(env_excel_path).expanduser()
        configured_excel_path = self.config.get("excel_path", DEFAULT_EXCEL_PATH)
        return Path(str(configured_excel_path)).expanduser()

    def _read_secret(self, *, env_name: str, label: str) -> str:
        """Read required secret from environment."""
        value = os.environ.get(env_name)
        if value:
            return value
        raise ValueError(f"{label} was not provided. Set environment variable {env_name}.")

    def _read_worksheet_entries(self) -> tuple[list[WorksheetEntry], list[dict[str, str]]]:
        """Read rows from worksheet and map them to trade inputs."""
        workbook = self.workbook_loader(self.excel_path, data_only=True, read_only=True)
        try:
            try:
                worksheet = workbook[self.worksheet_name]
            except KeyError as exc:
                raise ValueError(
                    f"Worksheet '{self.worksheet_name}' was not found in {self.excel_path}.",
                ) from exc
            header_row = self._locate_header_row(worksheet)
            entries: list[WorksheetEntry] = []
            skipped: list[dict[str, str]] = []
            for row_index in range(header_row + 1, worksheet.max_row + 1):
                pay_date_value = worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Pay Date"]).value
                if pay_date_value in (None, ""):
                    continue
                try:
                    pay_date = self._normalize_excel_date(pay_date_value)
                    roc_percent = self._normalize_decimal_string(
                        worksheet.cell(row=row_index, column=REQUIRED_HEADERS["ROC%"]).value,
                    )
                    roc_amount = self._normalize_decimal_string(
                        worksheet.cell(row=row_index, column=REQUIRED_HEADERS["ROC $"]).value,
                    )
                    gross_amount = self._normalize_decimal_string(
                        worksheet.cell(row=row_index, column=REQUIRED_HEADERS["Gross Amt"]).value,
                    )
                    exchange_rate = self._normalize_decimal_string(
                        worksheet.cell(row=row_index, column=EXCHANGE_RATE_COLUMN).value,
                    )
                except ValueError as exc:
                    skipped.append({"row": str(row_index), "reason": str(exc)})
                    continue
                entries.append(
                    WorksheetEntry(
                        row_index=row_index,
                        pay_date=pay_date,
                        roc_percent=roc_percent,
                        roc_amount=roc_amount,
                        gross_amount=gross_amount,
                        exchange_rate=exchange_rate,
                    ),
                )
        finally:
            workbook.close()
        return entries, skipped

    def _build_create_payload(self, portfolio_id: int, entry: WorksheetEntry) -> dict[str, Any]:
        """Build POST /trades create payload."""
        unique_identifier = (
            f"{self.unique_identifier_prefix}-{entry.pay_date.isoformat()}-{entry.roc_amount}-{entry.gross_amount}"
        )[:255]
        return {
            "trade": {
                "portfolio_id": portfolio_id,
                "holding_id": self.holding_id,
                "unique_identifier": unique_identifier,
                "transaction_date": entry.pay_date.strftime(API_DATE_FORMAT),
                "transaction_type": self.transaction_type,
                "price": 0.0,
                "capital_return_value": float(entry.roc_amount),
                "paid_on": entry.pay_date.strftime(API_DATE_FORMAT),
                "exchange_rate": float(entry.exchange_rate),
                "comments": entry.comment_text(),
                "state": self.created_state,
            },
        }

    def _build_update_payload(self, portfolio_id: int, entry: WorksheetEntry) -> dict[str, Any]:
        """Build PUT /trades/:id.json payload."""
        unique_identifier = (
            f"{self.unique_identifier_prefix}-{entry.pay_date.isoformat()}-{entry.roc_amount}-{entry.gross_amount}"
        )[:255]
        return {
            "trade": {
                "unique_identifier": unique_identifier,
                "transaction_type": self.transaction_type,
                "transaction_date": entry.pay_date.strftime(API_DATE_FORMAT),
                "portfolio_id": portfolio_id,
                "holding_id": self.holding_id,
                "price": 0.0,
                "capital_return_value": float(entry.roc_amount),
                "paid_on": entry.pay_date.strftime(API_DATE_FORMAT),
                "exchange_rate": float(entry.exchange_rate),
                "comments": entry.comment_text(),
            },
        }

    def _locate_header_row(self, worksheet: Worksheet) -> int:
        """Find header row containing all required header labels."""
        expected = set(REQUIRED_HEADERS)
        for row_index in range(1, worksheet.max_row + 1):
            values = {
                str(worksheet.cell(row=row_index, column=column_index).value).strip()
                for column_index in range(HEADER_START_COLUMN, max(REQUIRED_HEADERS.values()) + 1)
                if worksheet.cell(row=row_index, column=column_index).value is not None
                and str(worksheet.cell(row=row_index, column=column_index).value).strip()
            }
            if expected.issubset(values):
                return row_index
        raise ValueError(f"Worksheet '{self.worksheet_name}' is missing required headers: {', '.join(REQUIRED_HEADERS)}.")

    def _normalize_excel_date(self, value: Any) -> date:
        """Normalize an Excel value into a date."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, time):
            raise ValueError(f"Worksheet contains time-only value instead of date: {value}")
        try:
            return datetime.strptime(str(value).strip(), DATE_OUTPUT_FORMAT).date()
        except ValueError as exc:
            raise ValueError(f"Unable to parse date value: {value!r}") from exc

    def _normalize_decimal_string(self, value: Any) -> str:
        """Normalize numeric value to a simple string."""
        if value is None or str(value).strip() == "":
            return "0"
        text = str(value).strip()
        if self._is_excel_error_value(text):
            raise ValueError(f"Worksheet contains Excel error value: {text}")
        try:
            numeric = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError(f"Unable to parse numeric worksheet value: {value!r}") from exc
        output = format(numeric, "f")
        if "." in output:
            output = output.rstrip("0").rstrip(".")
        return output or "0"

    def _coerce_optional_bool(self, value: Any) -> bool | None:
        """Coerce optional boolean values from config or runtime."""
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().casefold()
        if text in {"true", "1", "yes", "y", "on"}:
            return True
        if text in {"false", "0", "no", "n", "off"}:
            return False
        raise ValueError(f"Unable to parse boolean value: {value!r}")

    def _is_excel_error_value(self, value: str) -> bool:
        """Return whether value looks like an Excel error literal."""
        return value in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NUM!", "#NAME?", "#NULL!"}

    def _to_decimal_or_none(self, value: str) -> Decimal | None:
        """Parse decimal string into Decimal for validation."""
        try:
            return Decimal(str(value).strip())
        except (InvalidOperation, ValueError):
            return None


class SharesightTradeApiClient:
    """urllib-backed Sharesight trade API client."""

    def __init__(self, *, api_base_url: str, token_url: str, client_id: str, client_secret: str) -> None:
        self.api_base_url = api_base_url.rstrip("/")
        self.token_url = token_url
        self.client_id = client_id
        self.client_secret = client_secret
        self._access_token: str | None = None

    def resolve_portfolio_id(self, portfolio_name: str) -> int:
        """Find portfolio ID by exact case-insensitive name."""
        response = self._request_json("GET", "/portfolios.json")
        portfolios = response.get("portfolios")
        if not isinstance(portfolios, list):
            raise RuntimeError("Unexpected Sharesight portfolios response: missing 'portfolios' list.")
        for item in portfolios:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name", ""))
            if name.casefold() == portfolio_name.casefold():
                return int(item["id"])
        raise RuntimeError(f"Unable to resolve portfolio {portfolio_name!r}.")

    def create_trade(self, payload: dict[str, Any]) -> TradeRecord:
        """POST create trade."""
        response = self._request_json("POST", "/trades.json", payload=payload)
        trade = response.get("trade", response)
        if not isinstance(trade, dict):
            raise RuntimeError("Unexpected response from create trade request.")
        return self._parse_trade_record(trade)

    def update_trade(self, trade_id: int, payload: dict[str, Any]) -> TradeRecord:
        """PUT update trade."""
        response = self._request_json("PUT", f"/trades/{trade_id}.json", payload=payload)
        trade = response.get("trade", response)
        if not isinstance(trade, dict):
            raise RuntimeError("Unexpected response from update trade request.")
        return self._parse_trade_record(trade)

    def delete_trade(self, trade_id: int) -> None:
        """DELETE trade."""
        self._request_json("DELETE", f"/trades/{trade_id}.json")

    def list_trades(
        self,
        portfolio_id: int,
        *,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[TradeRecord]:
        """GET trades for one portfolio."""
        params: dict[str, str] = {}
        if start_date is not None:
            params["start_date"] = start_date.isoformat()
        if end_date is not None:
            params["end_date"] = end_date.isoformat()
        query = f"?{parse.urlencode(params)}" if params else ""
        response = self._request_json("GET", f"/portfolios/{portfolio_id}/trades.json{query}")
        trades = response.get("trades")
        if not isinstance(trades, list):
            raise RuntimeError("Unexpected Sharesight trades response: missing 'trades' list.")
        return [self._parse_trade_record(item) for item in trades if isinstance(item, dict)]

    def close(self) -> None:
        """No-op hook for API resource cleanup."""
        return None

    def _parse_trade_roc(self, payload: dict[str, Any]) -> float | None:
        """Best-effort ROC / capital return from API payload."""
        for key in ("capital_return_value", "price"):
            raw_val = payload.get(key)
            if raw_val in (None, ""):
                continue
            try:
                return float(raw_val)
            except (TypeError, ValueError):
                continue
        return None

    def _parse_trade_fx(self, payload: dict[str, Any]) -> float | None:
        """Exchange rate from API payload."""
        raw_val = payload.get("exchange_rate")
        if raw_val in (None, ""):
            return None
        try:
            return float(raw_val)
        except (TypeError, ValueError):
            return None

    def _parse_trade_price_only(self, payload: dict[str, Any]) -> float | None:
        """Sharesight `price` field only (separate from capital_return_value / ROC)."""
        raw_val = payload.get("price")
        if raw_val in (None, ""):
            return None
        try:
            return float(raw_val)
        except (TypeError, ValueError):
            return None

    def _parse_trade_record(self, payload: dict[str, Any]) -> TradeRecord:
        """Extract fields from a trade JSON object."""
        trade_id = payload.get("id")
        company_event_id = payload.get("company_event_id")
        transaction_date = payload.get("transaction_date")
        transaction_type = str(payload.get("transaction_type", "")).upper()
        holding_id = payload.get("holding_id")
        unique_identifier = payload.get("unique_identifier")
        if company_event_id in (None, "") and isinstance(payload.get("company_event"), dict):
            company_event_id = payload["company_event"].get("id")
        return TradeRecord(
            id=None if trade_id in (None, "") else int(trade_id),
            company_event_id=None if company_event_id in (None, "") else int(company_event_id),
            transaction_date=(
                None
                if transaction_date in (None, "")
                else datetime.strptime(str(transaction_date).strip(), API_DATE_FORMAT).date()
            ),
            transaction_type=transaction_type,
            holding_id=None if holding_id in (None, "") else int(holding_id),
            unique_identifier=None if unique_identifier in (None, "") else str(unique_identifier),
            roc_value=self._parse_trade_roc(payload),
            exchange_rate_value=self._parse_trade_fx(payload),
            price_value=self._parse_trade_price_only(payload),
            raw=payload,
        )

    def _request_json(
        self,
        method: str,
        path_or_url: str,
        *,
        payload: dict[str, Any] | None = None,
        include_auth: bool = True,
    ) -> dict[str, Any]:
        """Send HTTP request and decode JSON response when present."""
        url = path_or_url if path_or_url.startswith("http") else f"{self.api_base_url}{path_or_url}"
        data = None
        headers = {"Accept": "application/json"}
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
            headers["Content-Type"] = "application/json"
        if include_auth:
            headers["Authorization"] = f"Bearer {self._get_access_token()}"

        req = request.Request(url, data=data, headers=headers, method=method)
        try:
            with request.urlopen(req) as response:
                body = response.read().decode("utf-8")
        except error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Sharesight API request failed: {method} {url} -> {exc.code}: {body}") from exc
        except error.URLError as exc:
            raise RuntimeError(f"Unable to reach Sharesight API at {url}: {exc.reason}") from exc
        if body.strip() == "":
            return {}
        decoded = json.loads(body)
        if not isinstance(decoded, dict):
            raise RuntimeError(f"Unexpected non-object JSON response from Sharesight API: {decoded!r}")
        return decoded

    def _get_access_token(self) -> str:
        """Fetch OAuth token lazily with client credentials."""
        if self._access_token is not None:
            return self._access_token
        body = parse.urlencode({"grant_type": "client_credentials"}).encode("utf-8")
        basic = b64encode(f"{self.client_id}:{self.client_secret}".encode("utf-8")).decode("ascii")
        headers = {
            "Accept": "application/json",
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
        }
        req = request.Request(self.token_url, data=body, headers=headers, method="POST")
        try:
            with request.urlopen(req) as response:
                raw = response.read().decode("utf-8")
        except error.HTTPError as exc:
            body_text = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"Sharesight OAuth token request failed: POST {self.token_url} -> {exc.code}: {body_text}",
            ) from exc
        except error.URLError as exc:
            raise RuntimeError(f"Unable to reach Sharesight OAuth endpoint at {self.token_url}: {exc.reason}") from exc
        decoded = json.loads(raw)
        if not isinstance(decoded, dict):
            raise RuntimeError(f"Unexpected token response from Sharesight OAuth endpoint: {decoded!r}")
        access_token = decoded.get("access_token")
        if not access_token:
            raise RuntimeError("Sharesight OAuth token response did not include 'access_token'.")
        self._access_token = str(access_token)
        return self._access_token


def format_currency(value: str) -> str:
    """Format a numeric string to 2dp without locale separators."""
    decimal = Decimal(value)
    return format(decimal.quantize(Decimal("0.01")), "f")
