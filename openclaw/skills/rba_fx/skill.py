"""RBA FX skill implementation."""

from __future__ import annotations

import csv
import os
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any, Callable

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from openclaw.skills._base.skill_base import SkillBase
from openclaw.skills.rba_fx.log_writer import ObsidianRunLogWriter

CSV_URL = "https://www.rba.gov.au/statistics/tables/csv/f11.1-data.csv"
EXCEL_PATH = "~/OneDrive/Documents/Finance/Personal CashFlow.xlsx"
EXCEL_PATH_ENV_VAR = "OPENCLAW_RBA_FX_EXCEL_PATH"
WORKSHEET_NAME = "FXRates"
DATE_COLUMN = "Date"
AUDUSD_COLUMN = "AUDUSD"
USDAUD_COLUMN = "USDAUD"
REQUEST_TIMEOUT_SECONDS = 30
DEFAULT_LOG_ENV = "local"
DEFAULT_LOG_OPERATOR = "unknown"
DEFAULT_OBSIDIAN_USER = "bobbyd"


@dataclass(slots=True)
class RunResult:
    """Summary of a completed RBA FX sync."""

    rows_downloaded: int
    rows_appended: int
    latest_sheet_date: str | None
    workbook_path: str
    worksheet_name: str


class RbaFxSkill(SkillBase):
    """Fetch RBA FX data and update the configured Excel worksheet."""

    def __init__(
        self,
        config_path: str | Path | None = None,
        *,
        csv_url: str | None = None,
        excel_path: str | Path | None = None,
        excel_path_env_var: str | None = None,
        worksheet_name: str | None = None,
        requests_get: Callable[..., requests.Response] | None = None,
        read_excel: Callable[..., pd.DataFrame] | None = None,
        workbook_loader: Callable[..., Any] | None = None,
    ) -> None:
        """Initialize the skill with optional dependency injection for testing."""
        super().__init__(config_path)
        self.csv_url = csv_url or str(self.config.get("csv_url", CSV_URL))
        self.excel_path_env_var = str(
            excel_path_env_var or self.config.get("excel_path_env_var", EXCEL_PATH_ENV_VAR),
        )
        self.excel_path = self._resolve_excel_path(excel_path)
        self.worksheet_name = str(
            worksheet_name or self.config.get("worksheet_name", WORKSHEET_NAME),
        )
        self.requests_get = requests_get or requests.get
        self.read_excel = read_excel or pd.read_excel
        self.workbook_loader = workbook_loader or load_workbook

    def _resolve_excel_path(self, excel_path: str | Path | None) -> Path:
        """Resolve the workbook path from explicit input, env var, or config."""
        if excel_path is not None:
            return Path(excel_path).expanduser()

        env_excel_path = os.getenv(self.excel_path_env_var)
        if env_excel_path:
            self.logger.info(
                "Using workbook path from environment variable %s.",
                self.excel_path_env_var,
            )
            return Path(env_excel_path).expanduser()

        configured_excel_path = self.config.get("excel_path", EXCEL_PATH)
        return Path(str(configured_excel_path)).expanduser()

    def _write_obsidian_run_log(self, result_payload: dict[str, Any]) -> None:
        """Write an Obsidian run log when explicitly enabled in config."""
        if not bool(self.config.get("obsidian_log_enabled", False)):
            return

        configured_logs_dir = self.config.get("obsidian_log_dir")
        if not configured_logs_dir:
            self.logger.warning(
                "Obsidian logging enabled, but `obsidian_log_dir` is not configured.",
            )
            return

        log_writer = ObsidianRunLogWriter(
            logs_dir=Path(str(configured_logs_dir)).expanduser(),
            operator=str(self.config.get("obsidian_log_operator", DEFAULT_LOG_OPERATOR)),
            environment=str(self.config.get("obsidian_log_environment", DEFAULT_LOG_ENV)),
            obsidian_user=str(self.config.get("obsidian_log_user", DEFAULT_OBSIDIAN_USER)),
            config_path=str(
                self.config.get("obsidian_log_config_path", "skills/rba_fx/config.yaml"),
            ),
            csv_url=self.csv_url,
            worksheet_name=self.worksheet_name,
            env_var_name=self.excel_path_env_var,
        )

        output_path = log_writer.write_log(
            result_payload,
            excel_path_resolved=str(self.excel_path),
            env_override_used=bool(os.getenv(self.excel_path_env_var)),
            command_used=str(
                self.config.get(
                    "obsidian_log_command_used",
                    "python - <<'PY' ... RbaFxSkill(...).run() ... PY",
                ),
            ),
            notes=["Generated automatically by RbaFxSkill."],
        )
        self.logger.info("Wrote Obsidian run log to %s.", output_path)

    def run(self) -> dict[str, Any]:
        """Download RBA FX data, append missing rows, and write the workbook."""
        self.logger.info("Starting RBA FX sync.")
        csv_text = self._download_csv()
        latest_rates = self._parse_rba_csv(csv_text)
        existing_sheet = self._read_existing_sheet()
        latest_sheet_date = self._get_latest_sheet_date(existing_sheet)
        _updated_sheet, missing_rows, rows_appended = self._merge_new_rows(existing_sheet, latest_rates)
        self._write_sheet(missing_rows)

        result = RunResult(
            rows_downloaded=len(latest_rates),
            rows_appended=rows_appended,
            latest_sheet_date=(
                latest_sheet_date.strftime("%d-%m-%Y")
                if latest_sheet_date is not None
                else None
            ),
            workbook_path=str(self.excel_path),
            worksheet_name=self.worksheet_name,
        )
        self.logger.info(
            "RBA FX sync completed with %s appended rows.",
            rows_appended,
        )
        result_payload = {
            "status": "success",
            "rows_downloaded": result.rows_downloaded,
            "rows_appended": result.rows_appended,
            "latest_sheet_date": result.latest_sheet_date,
            "workbook_path": result.workbook_path,
            "worksheet_name": result.worksheet_name,
        }
        self._write_obsidian_run_log(result_payload)
        return result_payload

    def _download_csv(self) -> str:
        """Download the RBA CSV source data."""
        self.logger.info("Downloading RBA CSV from %s.", self.csv_url)
        try:
            response = self.requests_get(self.csv_url, timeout=REQUEST_TIMEOUT_SECONDS)
            response.raise_for_status()
        except requests.RequestException as exc:
            self.logger.exception("Failed to download RBA CSV data.")
            raise RuntimeError("Unable to download RBA FX data.") from exc

        self.logger.info("Downloaded %s bytes of RBA CSV data.", len(response.text))
        return response.text

    def _parse_rba_csv(self, csv_text: str) -> pd.DataFrame:
        """Parse the RBA CSV into normalized FX rate rows."""
        self.logger.info("Parsing downloaded RBA CSV data.")
        csv_table_text = self._extract_csv_table_text(csv_text)
        raw_frame = pd.read_csv(StringIO(csv_table_text))
        date_source_column, rate_source_column = self._identify_source_columns(raw_frame)

        parsed = pd.DataFrame(
            {
                DATE_COLUMN: pd.to_datetime(
                    raw_frame[date_source_column],
                    dayfirst=True,
                    errors="coerce",
                ),
                AUDUSD_COLUMN: pd.to_numeric(
                    raw_frame[rate_source_column],
                    errors="coerce",
                ),
            },
        )

        malformed_mask = parsed[DATE_COLUMN].isna() | parsed[AUDUSD_COLUMN].isna()
        malformed_rows = int(malformed_mask.sum())
        if malformed_rows:
            self.logger.warning(
                "Skipping %s malformed RBA CSV rows.",
                malformed_rows,
            )

        parsed = parsed.loc[~malformed_mask].copy()
        parsed[USDAUD_COLUMN] = 1 / parsed[AUDUSD_COLUMN]
        parsed[DATE_COLUMN] = parsed[DATE_COLUMN].dt.strftime("%d-%m-%Y")
        parsed = parsed[[DATE_COLUMN, AUDUSD_COLUMN, USDAUD_COLUMN]]
        parsed = parsed.drop_duplicates(subset=[DATE_COLUMN]).sort_values(
            by=DATE_COLUMN,
            key=lambda series: pd.to_datetime(series, format="%d-%m-%Y", dayfirst=True),
        )
        parsed = parsed.reset_index(drop=True)

        self.logger.info("Parsed %s valid FX rows from RBA CSV.", len(parsed))
        return parsed

    def _extract_csv_table_text(self, csv_text: str) -> str:
        """Trim metadata rows and keep only the table header plus dated rows."""
        rows = list(csv.reader(StringIO(csv_text)))
        header_index = self._find_header_row_index(rows)
        header = rows[header_index]
        expected_width = len(header)
        data_rows = [header]

        for row in rows[header_index + 1 :]:
            if not row:
                continue
            first_cell = row[0].strip() if row else ""
            if not first_cell:
                continue
            parsed_date = pd.to_datetime(first_cell, dayfirst=True, errors="coerce")
            if pd.isna(parsed_date):
                continue

            normalized_row = list(row[:expected_width])
            if len(normalized_row) < expected_width:
                normalized_row.extend([""] * (expected_width - len(normalized_row)))
            data_rows.append(normalized_row)

        if len(data_rows) == 1:
            msg = "Unable to locate dated rows in RBA CSV."
            raise ValueError(msg)

        output = StringIO()
        csv.writer(output, lineterminator="\n").writerows(data_rows)
        return output.getvalue()

    def _find_header_row_index(self, rows: list[list[str]]) -> int:
        """Return the CSV row index containing the table headers."""
        for index, row in enumerate(rows):
            if not row:
                continue
            first_cell = row[0].strip().lstrip("\ufeff").lower()
            if first_cell in {"date", "title"}:
                return index

        msg = "Unable to locate the header row in RBA CSV."
        raise ValueError(msg)

    def _identify_source_columns(self, raw_frame: pd.DataFrame) -> tuple[str, str]:
        """Identify the date and AUD/USD source columns in the RBA CSV."""
        normalized_columns = {str(column).strip().lower(): str(column) for column in raw_frame.columns}
        date_column = next(
            (column for key, column in normalized_columns.items() if "date" in key),
            None,
        )
        rate_column = next(
            (
                column
                for key, column in normalized_columns.items()
                if "aud/usd" in key or "audusd" in key or "usd per aud" in key
            ),
            None,
        )

        if date_column is None:
            date_column = str(raw_frame.columns[0]) if len(raw_frame.columns) >= 1 else None
        if rate_column is None:
            numeric_candidates = [
                str(column)
                for column in raw_frame.columns
                if str(column) != date_column
            ]
            rate_column = numeric_candidates[0] if numeric_candidates else None

        if date_column is None or rate_column is None:
            msg = "Unable to identify required columns in RBA CSV."
            raise ValueError(msg)

        return date_column, rate_column

    def _read_existing_sheet(self) -> pd.DataFrame:
        """Load the configured worksheet from the target workbook."""
        self.logger.info(
            "Reading existing FX data from %s [%s].",
            self.excel_path,
            self.worksheet_name,
        )
        try:
            existing = self.read_excel(self.excel_path, sheet_name=self.worksheet_name)
        except ValueError as exc:
            self.logger.exception("Configured worksheet is missing.")
            raise ValueError(
                f"Worksheet '{self.worksheet_name}' was not found in {self.excel_path}.",
            ) from exc

        if DATE_COLUMN not in existing.columns:
            self.logger.info("Worksheet has no Date column yet; starting from an empty sheet.")
            return pd.DataFrame(columns=[DATE_COLUMN, AUDUSD_COLUMN, USDAUD_COLUMN])

        normalized = existing.copy()
        for column in (AUDUSD_COLUMN, USDAUD_COLUMN):
            if column not in normalized.columns:
                normalized[column] = pd.NA

        return normalized[[DATE_COLUMN, AUDUSD_COLUMN, USDAUD_COLUMN]]

    def _get_latest_sheet_date(self, existing_sheet: pd.DataFrame) -> pd.Timestamp | None:
        """Return the latest date currently stored in the worksheet."""
        if existing_sheet.empty:
            self.logger.info("Existing worksheet is empty.")
            return None

        parsed_dates = pd.to_datetime(
            existing_sheet[DATE_COLUMN],
            format="%d-%m-%Y",
            dayfirst=True,
            errors="coerce",
        )
        latest_date = parsed_dates.max()

        if pd.isna(latest_date):
            self.logger.info("Existing worksheet contains no valid dates.")
            return None

        self.logger.info("Latest date in worksheet is %s.", latest_date.strftime("%d-%m-%Y"))
        return latest_date

    def _merge_new_rows(
        self,
        existing_sheet: pd.DataFrame,
        latest_rates: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame, int]:
        """Append only rows newer than the latest worksheet date."""
        latest_sheet_date = self._get_latest_sheet_date(existing_sheet)

        if latest_sheet_date is None:
            missing_rows = latest_rates.copy()
        else:
            source_dates = pd.to_datetime(
                latest_rates[DATE_COLUMN],
                format="%d-%m-%Y",
                dayfirst=True,
            )
            missing_rows = latest_rates.loc[source_dates > latest_sheet_date].copy()

        rows_appended = len(missing_rows)
        self.logger.info("Identified %s new FX rows to append.", rows_appended)

        if existing_sheet.empty:
            updated_sheet = missing_rows
        else:
            updated_sheet = pd.concat([existing_sheet, missing_rows], ignore_index=True)

        updated_sheet = updated_sheet.drop_duplicates(subset=[DATE_COLUMN], keep="last")
        updated_sheet = updated_sheet.sort_values(
            by=DATE_COLUMN,
            key=lambda series: pd.to_datetime(series, format="%d-%m-%Y", dayfirst=True, errors="coerce"),
        ).reset_index(drop=True)
        return (
            updated_sheet[[DATE_COLUMN, AUDUSD_COLUMN, USDAUD_COLUMN]],
            missing_rows[[DATE_COLUMN, AUDUSD_COLUMN, USDAUD_COLUMN]],
            rows_appended,
        )

    def _write_sheet(self, missing_rows: pd.DataFrame) -> None:
        """Append missing rows and copy the previous row formatting."""
        if missing_rows.empty:
            self.logger.info("No new FX rows to write.")
            return

        self.logger.info("Appending %s rows to the workbook.", len(missing_rows))
        workbook = self.workbook_loader(self.excel_path)

        try:
            worksheet = workbook[self.worksheet_name]
        except KeyError as exc:
            workbook.close()
            self.logger.exception("Configured worksheet is missing during write.")
            raise ValueError(
                f"Worksheet '{self.worksheet_name}' was not found in {self.excel_path}.",
            ) from exc

        template_row_index = worksheet.max_row if worksheet.max_row > 1 else None

        for _, row in missing_rows.iterrows():
            worksheet.append(self._worksheet_row_values(row))
            if template_row_index is not None:
                self._copy_row_format(worksheet, template_row_index, worksheet.max_row)

        workbook.save(self.excel_path)
        workbook.close()
        self.logger.info("Workbook update saved to %s.", self.excel_path)

    def _worksheet_row_values(self, row: pd.Series) -> list[Any]:
        """Convert a normalized DataFrame row into worksheet cell values."""
        return [
            datetime.strptime(str(row[DATE_COLUMN]), "%d-%m-%Y").date(),
            float(row[AUDUSD_COLUMN]),
            float(row[USDAUD_COLUMN]),
        ]

    def _copy_row_format(
        self,
        worksheet: Worksheet,
        source_row_index: int,
        target_row_index: int,
    ) -> None:
        """Copy cell styles from one worksheet row to another."""
        for column_index in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=source_row_index, column=column_index)
            target_cell = worksheet.cell(row=target_row_index, column=column_index)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format

        source_dimensions = worksheet.row_dimensions[source_row_index]
        target_dimensions = worksheet.row_dimensions[target_row_index]
        target_dimensions.height = source_dimensions.height
        target_dimensions.hidden = source_dimensions.hidden
